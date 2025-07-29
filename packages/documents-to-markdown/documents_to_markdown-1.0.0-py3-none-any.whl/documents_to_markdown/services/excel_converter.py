#!/usr/bin/env python3
"""
Excel Document to Markdown Converter Service

This service handles conversion of Excel documents (.xlsx, .xlsm, .xls, .xlsb) to Markdown format.
"""

import sys
from pathlib import Path
from typing import List, Dict, Any, Optional
import re

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    print("Error: openpyxl is not installed. Please install it using:")
    print("pip install openpyxl")
    sys.exit(1)

try:
    import xlrd
except ImportError:
    print("Warning: xlrd is not installed. Legacy .xls files may not be supported.")
    print("Install it using: pip install xlrd")
    xlrd = None

from .base_converter import BaseDocumentConverter


class ExcelDocumentConverter(BaseDocumentConverter):
    """Converts Excel documents to Markdown format with full content preservation."""

    def get_supported_extensions(self) -> List[str]:
        """Get list of supported Excel document extensions."""
        extensions = ['.xlsx', '.xlsm', '.xlsb']
        if xlrd:
            extensions.append('.xls')
        return extensions
    
    def can_convert(self, file_path: Path) -> bool:
        """Check if this converter can handle Excel documents."""
        return file_path.suffix.lower() in self.get_supported_extensions()
    
    def _convert_document_to_markdown(self, doc_path: Path) -> str:
        """Convert an Excel document to Markdown format."""
        try:
            # Reset section counters for new document
            self._reset_section_counters()
            
            markdown_content = f"# {doc_path.stem}\n\n"
            
            # Handle different Excel formats
            if doc_path.suffix.lower() == '.xls' and xlrd:
                markdown_content += self._convert_xls_to_markdown(doc_path)
            else:
                markdown_content += self._convert_xlsx_to_markdown(doc_path)
            
            return markdown_content
            
        except Exception as e:
            self.logger.error(f"Error converting Excel document: {str(e)}")
            raise
    
    def _convert_xlsx_to_markdown(self, doc_path: Path) -> str:
        """Convert modern Excel formats (.xlsx, .xlsm, .xlsb) to Markdown."""
        try:
            workbook = load_workbook(doc_path, read_only=True, data_only=True)
            markdown_content = ""
            
            # Process each worksheet
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                sheet_content = self._convert_worksheet_to_markdown(worksheet, sheet_name)
                if sheet_content.strip():
                    markdown_content += sheet_content + "\n\n"
            
            workbook.close()
            return markdown_content
            
        except Exception as e:
            self.logger.error(f"Error processing Excel file: {str(e)}")
            raise
    
    def _convert_xls_to_markdown(self, doc_path: Path) -> str:
        """Convert legacy Excel format (.xls) to Markdown."""
        if not xlrd:
            raise ImportError("xlrd package is required for .xls files")
        
        try:
            workbook = xlrd.open_workbook(doc_path)
            markdown_content = ""
            
            # Process each worksheet
            for sheet_idx in range(workbook.nsheets):
                worksheet = workbook.sheet_by_index(sheet_idx)
                sheet_name = workbook.sheet_names()[sheet_idx]
                sheet_content = self._convert_xls_worksheet_to_markdown(worksheet, sheet_name)
                if sheet_content.strip():
                    markdown_content += sheet_content + "\n\n"
            
            return markdown_content
            
        except Exception as e:
            self.logger.error(f"Error processing XLS file: {str(e)}")
            raise
    
    def _convert_worksheet_to_markdown(self, worksheet: Worksheet, sheet_name: str) -> str:
        """Convert an openpyxl worksheet to Markdown format."""
        markdown_content = ""
        
        # Add worksheet heading with section numbering
        section_number = self._update_section_counter(2)
        markdown_content += f"## {section_number}{sheet_name}\n\n"
        
        # Get the data range
        if worksheet.max_row == 1 and worksheet.max_column == 1:
            # Empty worksheet
            markdown_content += "*This worksheet is empty.*\n\n"
            return markdown_content
        
        # Extract data from worksheet
        data = []
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                     min_col=1, max_col=worksheet.max_column, 
                                     values_only=True):
            # Convert None values to empty strings and handle different data types
            processed_row = []
            for cell in row:
                if cell is None:
                    processed_row.append("")
                else:
                    processed_row.append(str(cell))
            data.append(processed_row)
        
        # Remove completely empty rows from the end
        while data and all(cell == "" for cell in data[-1]):
            data.pop()
        
        if not data:
            markdown_content += "*This worksheet contains no data.*\n\n"
            return markdown_content
        
        # Convert to markdown table using Azure DevOps compatible format
        markdown_content += self._create_azure_devops_table(data)
        
        return markdown_content
    
    def _convert_xls_worksheet_to_markdown(self, worksheet, sheet_name: str) -> str:
        """Convert an xlrd worksheet to Markdown format."""
        markdown_content = ""
        
        # Add worksheet heading with section numbering
        section_number = self._update_section_counter(2)
        markdown_content += f"## {section_number}{sheet_name}\n\n"
        
        if worksheet.nrows == 0:
            markdown_content += "*This worksheet is empty.*\n\n"
            return markdown_content
        
        # Extract data from worksheet
        data = []
        for row_idx in range(worksheet.nrows):
            row = []
            for col_idx in range(worksheet.ncols):
                cell = worksheet.cell(row_idx, col_idx)
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    row.append("")
                elif cell.ctype == xlrd.XL_CELL_TEXT:
                    row.append(str(cell.value))
                elif cell.ctype == xlrd.XL_CELL_NUMBER:
                    # Handle integers vs floats
                    if cell.value == int(cell.value):
                        row.append(str(int(cell.value)))
                    else:
                        row.append(str(cell.value))
                elif cell.ctype == xlrd.XL_CELL_DATE:
                    # Convert Excel date to readable format
                    date_tuple = xlrd.xldate_as_tuple(cell.value, worksheet.book.datemode)
                    if date_tuple[:3] == (0, 0, 0):
                        # Time only
                        row.append(f"{date_tuple[3]:02d}:{date_tuple[4]:02d}:{date_tuple[5]:02d}")
                    else:
                        # Date and possibly time
                        row.append(f"{date_tuple[0]}-{date_tuple[1]:02d}-{date_tuple[2]:02d}")
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    row.append("TRUE" if cell.value else "FALSE")
                else:
                    row.append(str(cell.value))
            data.append(row)
        
        # Convert to markdown table using Azure DevOps compatible format
        markdown_content += self._create_azure_devops_table(data)
        
        return markdown_content
    


