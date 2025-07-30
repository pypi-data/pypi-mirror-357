def ifc2xlsx_main():
    print("Running ifc2xlsx GUI application...")
# (CC0) balaji.work
# Updated to have any parameters exported from IFC Models
# and always include GUID, Name, Revit Element ID, and Revit Family

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import ifcopenshell
import csv
import os
import sys
import pandas as pd
import webbrowser
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime  # For timestamp in Excel filename

# Log errors to a file when running as an exe
def log_errors_to_file(log_file="error_log.txt"):
    sys.stdout = open(log_file, "w")
    sys.stderr = sys.stdout

if getattr(sys, 'frozen', False):
    log_errors_to_file()

# Global variables
selected_properties = set()
available_properties = []
parameter_checkboxes = {}  # Initialize the dictionary of checkboxes

# Extract all parameters from the IFC files and save them to a text file
def extract_and_save_parameters(ifc_directory, output_txt_file):
    global available_properties
    all_columns = set()

    # Loop through each IFC file in the directory
    for root, _, files in os.walk(ifc_directory):
        for file_name in files:
            if file_name.lower().endswith(".ifc"):
                input_file_path = os.path.join(root, file_name)
                print(f"Processing file: {input_file_path}")
                ifc_file = ifcopenshell.open(input_file_path)
                elements = ifc_file.by_type("IfcElement")

                for element in elements:
                    if hasattr(element, "IsDefinedBy"):
                        for rel in element.IsDefinedBy:
                            if rel.is_a("IfcRelDefinesByProperties"):
                                prop_set = rel.RelatingPropertyDefinition
                                if hasattr(prop_set, "HasProperties"):
                                    for prop in prop_set.HasProperties:
                                        if hasattr(prop, "Name"):
                                            all_columns.add(prop.Name)
    # Save the available properties to a text file
    with open(output_txt_file, "w") as f:
        for param in sorted(all_columns):
            f.write(f"{param}\n")

    available_properties = sorted(all_columns)
    print(f"Saved {len(available_properties)} parameters to {output_txt_file}")


# Load selected properties from the checkbox list
def load_selected_properties_from_checkboxes():
    global selected_properties
    # Create a set of parameters where the checkbox was checked (var.get() is True)
    selected_properties = {param for param, var in parameter_checkboxes.items() if var.get()}
    print(f"Selected properties: {selected_properties}")


# Process all IFC files in a directory
def process_ifc_directory(input_dir, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    
    files_to_process = [f for root, _, files in os.walk(input_dir) for f in files if f.lower().endswith(".ifc")]
    total_files = len(files_to_process)
    
    if total_files == 0:
        print("No IFC files found.")
        return
    
    all_element_data = []
    all_columns = set(["GlobalId", "Name", "Type", "Revit_Element_ID", "Revit_Family"])  # default columns

    # Loop through each file and extract data
    for root, _, files in os.walk(input_dir):
        for file_name in files:
            if file_name.lower().endswith(".ifc"):
                input_file_path = os.path.join(root, file_name)
                print(f"Processing file: {input_file_path}")
    
                element_data, columns = extract_ifc_properties(input_file_path)
                all_element_data.extend(element_data)
                all_columns.update(columns)
    
    # Now we process the data and write it to CSV and Excel
    create_combined_output(all_element_data, sorted(all_columns), output_dir)


# Extract properties from IFC file
def extract_ifc_properties(ifc_file_path):
    print(f"Processing: {ifc_file_path}")
    ifc_file = ifcopenshell.open(ifc_file_path)
    elements = ifc_file.by_type("IfcElement")
    
    element_data = []
    # Initialize default columns for each element
    default_columns = {"GlobalId", "Name", "Type", "Revit_Element_ID", "Revit_Family"}
    all_columns = set(default_columns)
    
    for element in elements:
        # Get basic IFC properties
        element_id = element.GlobalId
        element_name = element.Name if element.Name else "Unknown"
        element_type = element.is_a()
        # Attempt to get Revit-specific values (if available)
        revit_id = getattr(element, "Tag", "N/A")
        revit_family = getattr(element, "ObjectType", element_type)
    
        # Prepare initial dictionary with default columns
        properties = {
            "GlobalId": element_id,
            "Name": element_name,
            "Type": element_type,
            "Revit_Element_ID": revit_id,
            "Revit_Family": revit_family
        }
    
        # Update all_columns with default keys
        all_columns.update(properties.keys())
    
        # Extract additional properties from property sets if available
        if hasattr(element, "IsDefinedBy"):
            for rel in element.IsDefinedBy:
                if rel.is_a("IfcRelDefinesByProperties"):
                    prop_set = rel.RelatingPropertyDefinition
                    if hasattr(prop_set, "HasProperties"):
                        for prop in prop_set.HasProperties:
                            if hasattr(prop, "Name") and hasattr(prop, "NominalValue"):
                                prop_name = prop.Name
                                prop_value = prop.NominalValue.wrappedValue
                                # Do not override default keys
                                if prop_name not in properties:
                                    properties[prop_name] = prop_value
                                all_columns.add(prop_name)
    
        element_data.append(properties)
    
    return element_data, sorted(all_columns)


# Combine the data and create both CSV and Excel outputs
def create_combined_output(all_element_data, all_columns, output_dir):
    # Filter data by selected columns if any are chosen, but always include default columns
    default_keys = {"GlobalId", "Name", "Type", "Revit_Element_ID", "Revit_Family"}
    if selected_properties:
        all_columns = list(default_keys.union({col for col in all_columns if col in selected_properties}))
    
    csv_output_path = os.path.join(output_dir, "combined_output.csv")
    with open(csv_output_path, mode="w", newline="", encoding="utf-8") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=all_columns)
        writer.writeheader()
        for data in all_element_data:
            # Write only selected columns to the CSV
            filtered_data = {key: data.get(key, "") for key in all_columns}
            # Add a single quote before any value starting with "="
            for k, v in filtered_data.items():
                if isinstance(v, str) and v.startswith("="):
                    filtered_data[k] = "'" + v
            writer.writerow(filtered_data)
    
    print(f"Saved CSV: {csv_output_path}")
    
    # Create Excel output
    create_excel_output(all_element_data, all_columns, output_dir)


# Create the Excel output
def create_excel_output(all_element_data, all_columns, output_dir):
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    validation_file = os.path.join(output_dir, f"validation_output_{timestamp}.xlsx")
    print(f"Saving validation file to: {validation_file}")
    
    # Create dataframe for Excel
    final_df = pd.DataFrame(all_element_data)
    final_df = final_df[all_columns]
    # Prepend a single quote if any cell's string value starts with "="
    final_df = final_df.applymap(lambda x: "'" + x if isinstance(x, str) and x.startswith("=") else x)
    
    try:
        final_df.to_excel(validation_file, index=False)
        validate_excel(validation_file)
        print(f"Validation file created at: {validation_file}")
        messagebox.showinfo(
            "Processing Complete",
            f"Processing complete!\n\nClick 'OK' to open the validation file.",
        )
        webbrowser.open(f"file://{validation_file}")
    except Exception as e:
        print(f"Error creating Excel: {e}")
        messagebox.showerror("Error", f"Error creating Excel: {e}")


# Validate Excel and apply error highlighting
def validate_excel(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Get header row and map column names to their positions
    header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    
    print(f"Columns in Excel: {header_row}")
    
    # Check if all required columns exist in the Excel sheet
    required_columns = [
        "CCILevel1ParentLocationID", "CCILevel1ParentTypeID",
        "CCILevel2ParentLocationID", "CCILevel2ParentTypeID",
        "CCILocationID", "CCIMultiLevelLocationID", "CCIMultiLevelTypeID"
    ]
    
    missing_columns = [col for col in required_columns if col not in header_row]
    if missing_columns:
        print(f"Missing columns in Excel: {missing_columns}")
        return
    
    try:
        # Perform validation for each row based on the selected columns
        idx_CCILevel1ParentLocationID = header_row.index("CCILevel1ParentLocationID")
        idx_CCILevel1ParentTypeID = header_row.index("CCILevel1ParentTypeID")
        idx_CCILevel2ParentLocationID = header_row.index("CCILevel2ParentLocationID")
        idx_CCILevel2ParentTypeID = header_row.index("CCILevel2ParentTypeID")
        idx_CCILocationID = header_row.index("CCILocationID")
        idx_CCIMultiLevelLocationID = header_row.index("CCIMultiLevelLocationID")
        idx_CCIMultiLevelTypeID = header_row.index("CCIMultiLevelTypeID")
    except ValueError as e:
        print(f"Required columns not found in the sheet: {e}")
        return
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        CCILevel2ParentTypeID = str(row[idx_CCILevel2ParentTypeID].value)
        CCILevel1ParentTypeID = str(row[idx_CCILevel1ParentTypeID].value)
        expected_type_id = f"§{CCILevel2ParentTypeID}.{CCILevel1ParentTypeID}"
        if str(row[idx_CCIMultiLevelTypeID].value) != expected_type_id:
            row[idx_CCIMultiLevelTypeID].fill = red_fill
    
        CCILevel2ParentLocationID = str(row[idx_CCILevel2ParentLocationID].value)
        CCILevel1ParentLocationID = str(row[idx_CCILevel1ParentLocationID].value)
        CCILocationID = str(row[idx_CCILocationID].value)
        expected_location_id = f"+{CCILevel2ParentLocationID}.{CCILevel1ParentLocationID}.{CCILocationID}"
        if str(row[idx_CCIMultiLevelLocationID].value) != expected_location_id:
            row[idx_CCIMultiLevelLocationID].fill = red_fill
    
    wb.save(excel_path)


# Select input directory
def select_input_directory():
    input_dir = filedialog.askdirectory(title="Select Input Directory")
    input_dir_entry.delete(0, tk.END)
    input_dir_entry.insert(0, input_dir)


# Select output directory
def select_output_directory():
    output_dir = filedialog.askdirectory(title="Select Output Directory")
    output_dir_entry.delete(0, tk.END)
    output_dir_entry.insert(0, output_dir)


# Select parameter list file
def select_property_list_file():
    property_file_path = filedialog.askopenfilename(title="Select Property List File", filetypes=[("Text Files", "*.txt")])
    if property_file_path:
        load_selected_properties(property_file_path)
        property_file_label.config(text=f"Loaded properties from: {property_file_path}")


def load_selected_properties(property_file_path):
    global selected_properties
    with open(property_file_path, 'r') as f:
        # Read and strip each line to get the property names.
        properties = [line.strip() for line in f if line.strip()]
        selected_properties = set(properties)
    
    # Update checkbox states if they are already loaded.
    for param, var in parameter_checkboxes.items():
        var.set(param in selected_properties)
    
    print(f"Loaded selected properties from file: {selected_properties}")


# Start processing
def start_processing():
    load_selected_properties_from_checkboxes()
    
    input_dir = input_dir_entry.get()
    output_dir = output_dir_entry.get()
    
    if not input_dir or not output_dir:
        messagebox.showerror("Error", "Please select both input and output directories.")
        return
    
    if not selected_properties:
        messagebox.showerror("Error", "Please select at least one parameter to export.")
        return
    
    try:
        print("Starting IFC to CSV conversion...")
        progress_var.set(0)
        progress_label.config(text="Initializing file processing...")
        process_ifc_directory(input_dir, output_dir)
        messagebox.showinfo("Success", "Processing complete! CSV files and validation sheet saved.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")


# Create Tkinter app
app = tk.Tk()
app.title("IFC to CSV Converter - (CC) balaji.work")
app.geometry("600x700")

tk.Label(app, text="Created for Eastern Ring Road Project by Balaji Balagurusami babs@cowi.com for COWI A/S.", fg="blue").pack(pady=5)
tk.Label(app, text="License: Creative Commons 0 1.0 Universal", fg="blue").pack(pady=5)

def open_github():
    webbrowser.open_new("https://github.com/balajibalagurusami/python/")

github_label = tk.Label(app, text="GitHub Repo", fg="blue", cursor="hand2")
github_label.pack(pady=5)
github_label.bind("<Button-1>", lambda e: open_github())

# Input directory
tk.Label(app, text="Select Input Directory:").pack()
input_dir_entry = tk.Entry(app, width=60)
input_dir_entry.pack(pady=5)
tk.Button(app, text="Browse", command=select_input_directory).pack()

# Output directory
tk.Label(app, text="Select Output Directory:").pack(pady=5)
output_dir_entry = tk.Entry(app, width=60)
output_dir_entry.pack(pady=5)
tk.Button(app, text="Browse", command=select_output_directory).pack()

# Property list selection
tk.Button(app, text="Load Property List", command=select_property_list_file).pack(pady=5)
property_file_label = tk.Label(app, text="No property list loaded.")
property_file_label.pack(pady=5)

# Scrollable frame for parameters
scroll_frame = tk.Frame(app)
scroll_frame.pack(pady=10)

# Create canvas and scrollbar for scrolling
canvas = tk.Canvas(scroll_frame)
scrollbar = tk.Scrollbar(scroll_frame, orient="vertical", command=canvas.yview)
canvas.config(yscrollcommand=scrollbar.set)

scrollable_frame = tk.Frame(canvas)

# Create the window inside the canvas
canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

scrollbar.pack(side="right", fill="y")
canvas.pack(side="left", fill="both", expand=True)

def load_checkboxes():
    global parameter_checkboxes
    # Clear previous checkboxes
    for widget in scrollable_frame.winfo_children():
        widget.destroy()
    
    parameter_checkboxes = {}
    
    for param in available_properties:
        var = tk.BooleanVar()
        cb = tk.Checkbutton(scrollable_frame, text=param, variable=var)
        cb.pack(anchor="w")
        parameter_checkboxes[param] = var
    
    # Update scroll region
    scrollable_frame.update_idletasks()
    canvas.config(scrollregion=canvas.bbox("all"))
    
    # Enable the Start Processing button after loading checkboxes
    start_button.config(state="normal")

# "Get Available Parameters" Button
tk.Button(app, text="Get Available Parameters", 
          command=lambda: [extract_and_save_parameters(input_dir_entry.get(), "available_parameters.txt"), load_checkboxes()]).pack(pady=5)

# Start button and progress bar
start_button = tk.Button(app, text="Start Processing", command=start_processing, state="disabled")
start_button.pack(pady=10)
progress_var = tk.DoubleVar()
progress_bar = ttk.Progressbar(app, variable=progress_var, length=400)
progress_bar.pack(pady=5)
progress_label = tk.Label(app, text="")
progress_label.pack()

app.mainloop()