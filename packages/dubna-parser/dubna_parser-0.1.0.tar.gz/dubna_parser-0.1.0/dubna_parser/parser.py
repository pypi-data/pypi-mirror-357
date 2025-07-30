import os

import openpyxl
from openpyxl.cell import Cell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.merge import MergedCellRange
from openpyxl.worksheet.worksheet import Worksheet

from dubna_parser import consts
from dubna_parser.downloader import download_sheets
from dubna_parser.enums import WeekDay, Type
from dubna_parser.models import Group, Pair, AlternatingPair, SchedulePair, GroupPairs, Schedule
from dubna_parser.utils import join_dicts, clean_spaces, extract_rgb_key


class ScheduleParser:
    def __init__(self):
        self.save_folder = 'downloads'
        self.file_with_url = 'links'
        self.degrees = consts.degrees
        self.wb: Workbook | None = None
        self.ws: Worksheet | None = None
        self.specializations_with_groups: dict[str, list[Group]] = dict()
        self.pairs_of_groups: list[GroupPairs] = list()
        self.specializations: set[str] = set()
        self.groups: set[Group] = set()
        self.classrooms: set[str] = set()
        self.subjects: set[str] = set()
        self.teachers: set[str] = set()
        self.all_specializations: set[str] = set()
        self.groups_from_row: list[Group] = list()
        self.group_parse_row_idx: int = 4
        self.group_parse_col_idx: int = 1
        self.group_parse_col_width: int = 2
        self.merged_pairs: dict[str, list[Group]] = dict()

    def set_default(self):
        self.wb = None
        self.ws = None
        self.specializations_with_groups: dict[str, list[Group]] = dict()
        self.pairs_of_groups: list[GroupPairs] = list()
        self.specializations: set[str] = set()
        self.groups: set[Group] = set()
        self.classrooms: set[str] = set()
        self.subjects: set[str] = set()
        self.teachers: set[str] = set()
        self.all_specializations: set[str] = set()
        self.groups_from_row: list[Group] = list()
        self.merged_pairs = dict()

    # методы для моделей
    def get_model_pair(self, classroom: str, subject: str, teacher: str, type_: Type) -> Pair:
        self.classrooms.add(classroom)
        self.subjects.add(subject)
        self.teachers.add(teacher)
        return Pair(classroom=classroom, subject=subject, teacher=teacher, type_=type_)

    def add_merged_pair(self, value: str, crange: MergedCellRange) -> None:
        merge_min_col_idx, merge_max_col_idx = (crange.min_col - self.group_parse_col_idx * self.group_parse_col_width,
                                                crange.max_col - self.group_parse_col_idx * self.group_parse_col_width)
        merge_groups = self.groups_from_row[merge_min_col_idx-1:merge_max_col_idx]
        self.merged_pairs[value] = merge_groups

    def serialize_group(self, group, specialization: str) -> Group:
        group_number = 0
        if group.isdigit():
            group_number = int(group)
        group_model = Group(group=group_number, specialization=specialization)
        self.groups.add(group_model)
        self.specializations.add(specialization)
        return group_model

    def download(self, file_with_url: str, save_folder: str):
        self.file_with_url, self.save_folder = file_with_url, save_folder
        download_sheets(file_with_url, save_folder)

    def get_not_empty_columns(self, index_start=3, row_start=4) -> (int, int):
        not_empty_column_index = index_start
        for index in range(index_start, self.ws.max_column + 1):
            if self.ws.cell(row=row_start, column=not_empty_column_index).value is None:
                break
            not_empty_column_index += 1
        return index_start, not_empty_column_index

    def get_specializations_from_row(self, index_start=3,
                                     index_end=3, index_row=4) -> dict[str, list[Group]]:
        groups: dict[str, list[Group]] = dict()
        self.groups_from_row = list()
        for col in self.ws.iter_cols(min_col=index_start, max_col=index_end,
                                     min_row=index_row, max_row=index_row,
                                     values_only=True):
            for cell in col:
                if cell:
                    group_with_specialization = cell.strip().split(' ', 1)
                    if len(group_with_specialization) == 1:
                        number_group, special = cell.strip().split('(', 1)
                        group_with_specialization = number_group, special
                    specialization = group_with_specialization[1].strip("()")
                    group = group_with_specialization[0]

                    group = self.serialize_group(group, specialization)
                    self.groups_from_row.append(group)
                    if specialization in groups:
                        groups[specialization].append(group)
                    else:
                        new_group = list()
                        new_group.append(group)
                        groups[specialization] = new_group
        return groups

    def get_indexes_of_weeks(self, column_index=1):
        day_indices = dict()
        for crange in self.ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = crange.bounds
            if min_col == column_index and max_col == column_index:
                day_indices[self.ws.cell(row=min_row, column=min_col).value] = (min_row, max_row)
        return day_indices

    def get_pair_from_row(self, row_cell: Cell, row_value: str | None) -> Pair:
        if row_value is None:
            row = clean_spaces(row_cell.value.replace('/', ''))
        else:
            row = row_value
        for degree in self.degrees:
            row = row.replace(degree, '').strip()
        classroom = ''
        teacher = ''
        for letter in row:
            if letter != ' ':
                classroom += letter
            else:
                break
        space_count = 0
        rev_row = row[::-1]
        for letter in rev_row:
            if space_count >= 2:
                break
            if letter != ' ':
                teacher += letter
            else:
                teacher += ' '
                space_count += 1
        teacher = teacher[::-1].strip()
        subject = clean_spaces(row.replace(teacher, '').replace(classroom, ''))
        return self.get_model_pair(classroom, subject, teacher,
                                   consts.color_to_type.get(extract_rgb_key(row_cell), Type.SEMINAR))

    def get_rows(self, column_index, first_row_index, second_row_index) -> tuple[Cell, Cell]:
        row1, row2 = '', ''
        for crange in self.ws.merged_cells.ranges:
            if row1 and row2:
                break
            min_col, min_row, max_col, max_row = crange.bounds
            if min_col <= column_index <= max_col and min_row == first_row_index:
                row1 = self.ws.cell(row=min_row, column=min_col)
                self.add_merged_pair(row1.value, crange)
            if min_col <= column_index <= max_col and min_row == second_row_index:
                row2 = self.ws.cell(row=min_row, column=min_col)
                self.add_merged_pair(row2.value, crange)
        else:
            if row1:
                row2 = self.ws.cell(row=second_row_index, column=column_index)
            elif row2:
                row1 = self.ws.cell(row=first_row_index, column=column_index)
            else:
                row1 = self.ws.cell(row=first_row_index, column=column_index)
                row2 = self.ws.cell(row=second_row_index, column=column_index)
        return row1, row2

    def get_single_row(self, column_index, row_index) -> Cell:
        row = ''
        for crange in self.ws.merged_cells.ranges:
            if row:
                break
            min_col, min_row, max_col, max_row = crange.bounds
            if min_col <= column_index <= max_col and min_row == row_index:
                row = self.ws.cell(row=min_row, column=min_col)
                self.add_merged_pair(row, crange)
        else:
            row = self.ws.cell(row=row_index, column=column_index)
        return row

    def get_pair(self, row_cell_1: Cell, row_cell_2: Cell) -> Pair | AlternatingPair | None:
        pair = None
        row1, row2 = row_cell_1.value, row_cell_2.value
        if row1 == 'с/к Олимп':
            classroom = row1
            subject = row2
            teacher = 'undefined'
            pair = Pair(classroom=classroom, subject=subject, teacher=teacher, type_=Type.PHYSICAL)
        elif row1 and row2:
            if row1.count('/') >= 1 and row2.count('/') >= 1:
                pair1 = self.get_pair_from_row(row_cell_1, row1)
                pair2 = self.get_pair_from_row(row_cell_2, row2)
                pair = AlternatingPair(odd_week=pair1, even_week=pair2)
                # пара мигалка
            else:
                row = row1 + '  ' + row2
                pair = self.get_pair_from_row(row_cell_1, row)
        elif row1:
            pair1 = self.get_pair_from_row(row_cell_1, row1)
            if row1.count('/') == 0:
                pair = pair1
            if row1.count('/') > 0:
                pair = AlternatingPair(odd_week=pair1, even_week=None)
        elif row2:
            pair2 = self.get_pair_from_row(row_cell_2, row2)
            if row2.count('/') == 0:
                pair = pair2
            if row2.count('/') > 0:
                pair = AlternatingPair(odd_week=None, even_week=pair)
        else:
            pair = None
        return pair

    def get_single_pair(self, row_cell: Cell) -> Pair | AlternatingPair | None:
        pair = None
        row = row_cell.value
        if row:
            row.strip()
            if row.count('/') == 0:
                pair = self.get_pair_from_row(row_cell, row)
            else:
                if row[0] == '/':
                    pair2 = self.get_pair_from_row(row_cell, row)
                    pair = AlternatingPair(odd_week=None, even_week=pair2)
                elif row[-1] == '/':
                    pair1 = self.get_pair_from_row(row_cell, row)
                    pair = AlternatingPair(odd_week=pair1, even_week=None)
                else:
                    print("тут уже мне неизвестно как парсить...")
        return pair


    def get_pairs_for_group(self, indices_of_week, column_index: int) -> dict[WeekDay, list[SchedulePair]]:
        group_pairs: dict[WeekDay, list[SchedulePair]] = dict()
        expected_day = WeekDay('вторник')
        for week_day, day_index in indices_of_week.items():
            start_day_index, end_day_index = day_index
            pairs_for_day: list[SchedulePair] = list()
            for i in range(start_day_index, end_day_index, 2):
                # перебираем все пары длинной в две строки
                first_row_index, second_row_index = i, i + 1
                row1, row2 = self.get_rows(column_index, first_row_index, second_row_index)
                pair_number = (i - start_day_index) // 2 + 1
                pair = self.get_pair(row1, row2)
                schedule_pair = SchedulePair(pair_number=pair_number, pair=pair)
                pairs_for_day.append(schedule_pair)

            if (end_day_index - start_day_index) % 2 == 1:
                # осталась последняя пара в одну строку
                row = self.get_single_row(column_index, end_day_index)
                pair_number = (end_day_index - start_day_index) // 2 + 1
                pair = self.get_single_pair(row)
                pair = SchedulePair(pair_number=pair_number, pair=pair)
                pairs_for_day.append(pair)
            if week_day:
                group_pairs[WeekDay(week_day)] = pairs_for_day
            else:
                group_pairs[expected_day] = pairs_for_day
        return group_pairs

    def parse(self, save_folder: str, get_set=False) -> Schedule:
        """перед парсингом не открывать таблицы! Ничего не трогать!"""
        self.set_default()
        all_files = os.listdir(save_folder)
        for filename in all_files:
            current_file = os.path.join(save_folder, filename)
            excel_file: Workbook = openpyxl.load_workbook(current_file)
            specializations_with_groups_from_file = dict()
            for name in excel_file.sheetnames:
                self.ws: Worksheet = excel_file[name]
                indices_of_week = self.get_indexes_of_weeks()
                start_index, end_index = self.get_not_empty_columns()
                specializations = self.get_specializations_from_row(start_index, end_index)
                all_values = [item for sublist in specializations.values() for item in sublist]
                for group_index, group in zip(range(start_index, end_index), all_values):
                    pairs_for_group = self.get_pairs_for_group(indices_of_week, group_index)
                    try:
                        pair_of_group = GroupPairs(group=group, group_pairs=pairs_for_group)
                        self.pairs_of_groups.append(pair_of_group)
                    except Exception as e:
                        print(e)
                        print("Данные: ")
                        print(group)
                        print(pairs_for_group)

                specializations_with_groups_from_file = join_dicts(specializations_with_groups_from_file,
                                                                   specializations)
            self.specializations_with_groups = join_dicts(self.specializations_with_groups,
                                                          specializations_with_groups_from_file)
        if get_set:
            return Schedule(
                specializations_with_groups_from_file=self.specializations_with_groups,
                specializations=self.specializations,
                groups=self.groups,
                classrooms=self.classrooms,
                teachers=self.teachers,
                subjects=self.subjects,
                schedule_pairs=self.pairs_of_groups,
                merged_pairs=self.merged_pairs,
            )
        return Schedule(
            specializations_with_groups_from_file=self.specializations_with_groups,
            specializations=list(self.specializations),
            groups=list(self.groups),
            classrooms=list(self.classrooms),
            teachers=list(self.teachers),
            subjects=list(self.subjects),
            schedule_pairs=self.pairs_of_groups,
            merged_pairs=self.merged_pairs,
        )
