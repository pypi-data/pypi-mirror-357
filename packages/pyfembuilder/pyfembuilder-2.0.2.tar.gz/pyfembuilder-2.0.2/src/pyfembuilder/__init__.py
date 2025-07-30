import os
import sys
from shutil import copyfile

def version():
    return 'pyFEMStudio 2.0.0c (November 2023)'


def write(build_extension='.inp'):
    params=[]
    params_pass1=[] #this is new, we make a first pass to identify all variables, only the variables that do not contain any param from this list after the equal sign are considered params of the template
    max_line_parameters=0

    NumberOfHeadRows=3

    cwd = os.getcwd()

    sourcefiles=[]

    ask_to_delete=0

    minver=-1

    filename_suffix = ""

    files_to_open=[]

    just_check_template=False
    warnings=[]

    template_file='template.inp'

    for m in sys.argv:
        if 'checktemplate' in m.lower():
            just_check_template=True
        if 'ignoresuffix' in m.lower():
            filename_suffix=""
        elif 'minver' in m.lower():
            try:
                minver=int(m.split("=")[1])
                print("We will create only cases with version>"+str(minver))
            except:
                print("Could not understand the minver to use. Use argument minver=# to only create cases with version>#")
                minver=-1
        elif 'xlsx' in m.lower():
            print("We will open the file", m)
            files_to_open.append(m)
        if 'template' in m.lower():
            try:
                template_file=m.split("=")[1]
                print("We will use the template file=", template_file)
            except:
                print("Could not understand the template file argument. using template.inp")
                template_file='template.inp'

    if len(files_to_open)==0:
        print("We will open all *.xlsx files in current folder.")
        for file in os.listdir("./"):
            if file.lower().endswith(".xlsx") and not file.lower().startswith("~$"):
                print("-> ",file)
                files_to_open.append(file)
            elif file.lower().startswith("~$"):
                print("-- ignored: ", file)

    for file in os.listdir("./"):
        if file.lower().endswith(".mac") or file.lower().endswith(".inp"):
            sourcefiles.append(file)


    print("Reading template file",template_file,'.')
    f=open(template_file)
    template_lines=f.readlines()
    for j in range(0,len(template_lines)):
        if "=" in template_lines[j]:
            ignore=False
            if "!" in template_lines[j]:
                if template_lines[j][0]!="!":
                    #ignore line
                    ignore=False
                else:
                    ignore=True
            if not ignore:
                arg = template_lines[j].split("=")
                params_pass1.append([arg[0].lower(),j,template_file,-1,'unit'])
        if "end_of_parameters" in template_lines[j]:
            print("We stop reading parameters at line",j)
            max_line_parameters=j

    if max_line_parameters==0:
        max_line_parameters=len(template_lines)

    #params_pass1.append("\t")   #we add the tab to skip parameters with tabs
    #second pass to identify params
    for j in range(0,len(template_lines)):
        if "=" in template_lines[j]:
            ignore=False
            if "!" in template_lines[j]:
                if template_lines[j][0]!="!":
                    #ignore line
                    ignore=False
                else:
                    ignore=True
            if j>0:
                if "!ignorenextparameter" in template_lines[j-1].lower():
                    print("Line ignored", template_lines[j].replace("\n",""))
                    ignore=True
            if not ignore:
                arg = template_lines[j].split("=")
                arg[1]=arg[1].split("!")[0]
                arg[1]=arg[1].replace("+","").replace("\\","").replace("/","").replace("e","").replace("-","").replace("\n","").replace("\t","").replace(" ","").replace(".","")
                #print(arg[1], not arg[1].isdigit())

                contains=False
                if not arg[1].isdigit():
                    contains=True

                if '\'' in arg[1]:
                    contains=False

                # for m in params_pass1:
                #     if m[0]==str(arg[1].lower().replace("\n","").replace("\t","")):
                #         contains=True
                #         break
                if not contains and j<max_line_parameters:
                    for m in params:
                        if m[0]==str(arg[0].lower().replace("\n","").replace("\t","")):
                            warnings.append("WARNING: the parameter -> "+arg[0]+' <- defined at line '+str(m[1]+1)+' is redefined at line '+str(j+1)+'. Line '+str(j+1)+ ' is ignored.')
                            print(warnings[-1])
                            contains=True
                            break
                if not contains and j<max_line_parameters:
                    params.append([arg[0].lower(),j,template_file,-1,'unit'])
                    print([arg[0].lower(),j,template_file,-1,'unit'])
                elif not contains and j>max_line_parameters:
                    #this does not appear at the end
                    print("WARNING: parameter ->"+arg[0]+' <- appears after enf_of_parameters instruction and was ignored')
                    #warnings.append("WARNING: parameter ->"+arg[0]+' <- appears after enf_of_parameters instruction and was ignored')
                    #print(warnings[-1])


    f.close()
    print("Reading template is OK.")
    if just_check_template:
        print("\nWe are leaving because we are just checking template parameters.")
        exit()

    #for v in params:
    #    print(v)
    #    #input()

    col_of_Folder=0
    col_of_Filename=0
    col_of_version=0
    col_of_output=0

    num_cases_created=0
    num_cases_ignored=0

    for file in files_to_open:
        print("\nOpening file ", file)
        from openpyxl import load_workbook
        wb = load_workbook(filename=file,data_only=True)
        w=wb.sheetnames[0]
        for i in range(1,wb[w].max_column+1):
            #print(str(parameters_to_read[param]).lower())
            test_v=wb[w].cell(1,i).value.lower()
            param_found=False
            if test_v=="folder":
                col_of_Folder = i
                param_found=True
            if test_v=="filename":
                param_found=True
                col_of_Filename = i
            if test_v=="version":
                param_found=True
                col_of_version = i
            if test_v=="outputfile":
                #after we find the outputfile column every parameter to the right is ignored!
                break


            for j in range(0,len(params)):
                if params[j][0]==test_v:
                    param_found=True
                    params[j][3]=i
                    params[j][4]=test_v=wb[w].cell(2,i).value.lower()    #important to have the units, use unit [str] to put the value enclosed in ''
            if not param_found:
                warnings.append("WARNING: the parameter "+str(test_v)+" was defined in the Excel file but not found on the template.")



        totalcases=wb[w].max_row-NumberOfHeadRows
        print("TotalCases=",totalcases)

        for line_of_case in range(1+NumberOfHeadRows, totalcases+1+NumberOfHeadRows):
            case_version=int(wb[w].cell(line_of_case,col_of_version).value)
            if case_version>=minver:
                Folder = wb[w].cell(line_of_case,col_of_Folder).value
                Filename = str(wb[w].cell(line_of_case,col_of_Filename).value)
                Filename = Filename.lower()+filename_suffix
                #print(f'The case: {Folder}/{Filename}.inp')
                print(f'{line_of_case-NumberOfHeadRows}/{totalcases}')

                #we create the cases here!
                if os.path.isdir(Folder) == False:
                    print("Creating dir...")
                    os.makedirs(Folder)
                #print("Dir exists!")
                #os.chdir(Folder)
                #check if file exists


                for v in params:
                    if v[3]!=-1:
                        the_value=wb[w].cell(line_of_case,v[3]).value
                        if v[4]!='[str]':
                            template_lines[v[1]]=f'{v[0]}={the_value}    !{v[4]}\n'
                        else:
                            template_lines[v[1]]=f"{v[0]}='{str(the_value).lower()}'\n"
                        #print(v)
                #checks if file exist and if version is higher or lower
                newer=True
                if os.path.exists(cwd+"/"+Folder+"/"+Filename+build_extension):
                    #
                    f=open(cwd+"/"+Folder+"/"+Filename+build_extension)
                    version_line=f.readline().replace("\n","")
                    print("line: "+version_line)
                    if 'version' in version_line:
                        arg=version_line.split("=")[1]

                        if case_version > int(arg):
                            newer=True
                            if os.path.exists(cwd+"/"+Folder+"/"+Filename+'.out') and ask_to_delete==0:
                                print("File exists:\n"+cwd+"/"+Folder+"/"+Filename+'.out')
                                print("Old output files were found. Do you wish to remove them (y/n):")
                                question=input()
                                if question=='y' or question=='Y':
                                    ask_to_delete=1
                                else:
                                    ask_to_delete=-1
                            if os.path.exists(cwd+"/"+Folder+"/"+Filename+'.out') and ask_to_delete==1:
                                #we remove the old outputs
                                os.remove(cwd+"/"+Folder+"/"+Filename+'.out')
                                print("Deleted " + Folder+"/"+Filename+'.out')
                        else:
                            newer=False
                    f.close()
                if newer:
                    for file_to_copy in sourcefiles:
                        if not 'template' in file_to_copy:
            #                copyfile(cwd + "/"+file_to_copy,cwd+"/"+Folder+"/"+Filename+'.inp')
            #            else:
                            #copy all the files except the template, to be changed by params...
                            copyfile(cwd + "/"+file_to_copy,cwd+"/"+Folder+"/"+file_to_copy)

                    fn = open(cwd+"/"+Folder+"/"+Filename+build_extension,"w")
                    print(f"!version={int(wb[w].cell(line_of_case,col_of_version).value)}", file=fn)
                    fn.writelines(template_lines)
                    fn.close()
                    num_cases_created+=1
                else:
                    num_cases_ignored+=1
                    print('same version, ignored.')
                #close and exit!
                #input()
        wb.close()

        if len(warnings)>0:
            print("\n\nDone with ",len(warnings),'warnings issued:\n')
            for m in warnings:
                print(m)
            print("\nConsider adding end_of_parameters in template after the complete definition of parameters or !ignorenextparameter to avoid non problematic warnings.")
            print("To check the template without creating the cases add 'checktemplate' to the call to the builder file")

        if num_cases_created>0 and num_cases_ignored>0:
            print("\nDone. We created",num_cases_created,"cases and ignored",num_cases_ignored,"cases.")
        if num_cases_created>0 and num_cases_ignored==0:
            print("\nDone. We created",num_cases_created,"cases. No cases ignored.")
        if num_cases_created==0 and num_cases_ignored>0:
            print("\nDone. No cases were created because",num_cases_ignored,"cases were ignored.")





    #    #input()


#0_read_difftestload_v4.py
def read(testdispload=0):
    params=[]
    cases=[]

    NumberOfHeadRows=3

    cwd = os.getcwd()

    col_of_Folder=0
    col_of_Filename=0
    col_of_reason=0
    col_of_testloaddisp=0
    col_of_maxdisp=0
    col_of_outputfile=0
    col_of_version=0

    filename_suffix = ""

    files_to_open=[]

    minver=-1

    for m in sys.argv:
        if 'ignoresuffix' in m.lower():
            print("Ignoring the suffix")
            filename_suffix=""
        elif 'xlsx' in m.lower():
            print("We will read the file", m)
            files_to_open.append(m)
        elif 'minver' in m.lower():
            try:
                minver=int(m.split("=")[1])
                print("We will read only cases with version>"+str(minver))
            except:
                print("Could not understand the minver to use. Use argument minver=# to only read cases with version>#")
                minver=-1

    if len(files_to_open)==0:
        print("We will read all *.xlsx files in current folder.")
        for file in os.listdir("./"):
            if file.lower().endswith(".xlsx") and not file.lower().startswith("~$"):
                print("-> ",file)
                files_to_open.append(file)
            elif file.lower().startswith("~$"):
                print("-- ignored: ", file)

    for file in files_to_open:
        print("\nOpening file ", file)
        #copyfile(file,file+".backup")      #we no longer make backups before reading
        from openpyxl import load_workbook
        wb = load_workbook(filename=file, data_only=True)
        w=wb.sheetnames[0]
        for i in range(1,wb[w].max_column+1):
            #print(str(parameters_to_read[param]).lower())
            test_v=wb[w].cell(1,i).value.lower()
            unit_info=wb[w].cell(2,i).value.lower()
            if test_v=="folder": col_of_Folder = i
            if test_v=="filename": col_of_Filename = i
            if test_v=="reasonfortermination": col_of_reason=i
            if test_v=="outputfile": col_of_outputfile=i
            if test_v=="version": col_of_version = i
            if test_v=="testmaxdisp":
                col_of_maxdisp=i
                print('testmaxdisp is at',i)
            if test_v=="testloaddisp":
                col_of_testloaddisp=i
                print('testloaddisp',i)
        for i in range(1,wb[w].max_column+1):
            #print(str(parameters_to_read[param]).lower())
            test_v=wb[w].cell(1,i).value.lower()
            unit_info=wb[w].cell(2,i).value.lower()
            if i >= col_of_outputfile: params.append([test_v,i,unit_info])

        if col_of_reason==0:
            print("ERROR! The Excel with files and folders must contain a column (ReasonForTermination)")
            break

        totalcases=wb[w].max_row-NumberOfHeadRows
        print("TotalCases=",totalcases)

        for line_of_case in range(1+NumberOfHeadRows, totalcases+1+NumberOfHeadRows):
            Folder = wb[w].cell(line_of_case,col_of_Folder).value
            Filename = wb[w].cell(line_of_case,col_of_Filename).value
            Filename = Filename.lower()
            cases.append([Folder,Filename])

        wb.close()
        wb = load_workbook(filename=file)    #reopens for latter save...
        w=wb.sheetnames[0]
        for line_of_case in range(1+NumberOfHeadRows, totalcases+1+NumberOfHeadRows):
            #print(f'{line_of_case-NumberOfHeadRows}/{totalcases}')
            case_version=int(str(wb[w].cell(line_of_case,col_of_version).value).replace("=",""))

            if case_version < minver:
                continue

            reason = "Error reading file"
            Folder=cases[line_of_case-NumberOfHeadRows-1][0]
            Filename=cases[line_of_case-NumberOfHeadRows-1][1]+filename_suffix
            thefile = cwd+"/"+Folder+"/"+Filename
            #try:
            print(f'{line_of_case-NumberOfHeadRows}/{totalcases}')
            print(f'The case: {Folder}/{Filename}.out')

            if os.path.exists(thefile+'.out'):
                #print(f'The case: {Folder}/{Filename}.out')

                f=open(thefile+".out")
                alllines=f.readlines()

                #print(len(alllines))
                for j in range(len(alllines)-1,0,-1):
                    if alllines[j].startswith(" REASON FOR TERMINATION"):
                        reason=alllines[j][42:]
                        reason=reason.replace("\n","")
                        wb[w].cell(line_of_case,col_of_reason).value=reason
                        #print(reason)
                f.close()

            #except:
            #    print('Error reading file.')
            if os.path.exists(thefile+'_gmnia.res.csv'):
                print("opening gmnia file!")
                thefile+='_gmnia'


            print(thefile+'.csv')
            if col_of_testloaddisp>0 and os.path.exists(thefile+'.csv') and testdispload==0:
                #print("yo!")
                mmax=0
                mmin=0
                valor=0
                f=open(thefile+'.csv')
                #print("all ok!")
                alllines=f.readlines()
                if len(alllines)>10:
                    for j in range(10,len(alllines)):
                        arg=alllines[j].split(";")
                        valor=abs(float(arg[0].replace("\n","")))
                        mmax=max(mmax,valor)
                    print(mmax)
                    #print(alllines[len(alllines)-1])
                    arg=alllines[len(alllines)-1].split(";")
                    mmin=abs(float(arg[-1].replace("\n","")))
                    if mmin==0: mmin=-1
                    valor=mmin/mmax
                    wb[w].cell(line_of_case,col_of_testloaddisp).value=valor
                f.close()

            if col_of_testloaddisp>0 and os.path.exists(thefile+'.csv') and testdispload==1:
                f=open(thefile+'.csv')
                #print("all ok!")
                alllines=f.readlines()
                arg=alllines[len(alllines)-1].replace('\n','').split(";")
                #print(arg)
                max_time=float(arg[0])
                threshold=0.99
                test_coord=3
                last_step=float(arg[test_coord])
                first_step=0
                for j in range(len(alllines)-2,0,-1):
                    arg=alllines[j].replace('\n','').split(";")
                    cur_time=float(arg[0])
                    cur_test=cur_time/max_time
                    if cur_test>threshold:
                        #print(j,'conta',cur_time,max_time,cur_time/max_time)
                        first_step=float(arg[test_coord])
                    else:
                        pass

                if first_step==0: first_step=last_step

                wb[w].cell(line_of_case,col_of_testloaddisp).value=last_step/first_step
                if col_of_maxdisp>0:
                    wb[w].cell(line_of_case,col_of_maxdisp).value=last_step
                    print('last_step',last_step)

                #print(first_step,last_step,)

            if os.path.exists(thefile+'.res.csv'):
                f=open(thefile+".res.csv")
                alllines=f.readlines()
                if len(alllines)>1:
                    all_arg=alllines[0].replace('\t','').replace(' ','').replace('\n','').split(";")
                    all_arg2=alllines[1].replace('\t','').replace(' ','').replace('\n','').split(";")
                    z=0
                    for arg in all_arg:
                        for v in params:
                            if v[0].lower().replace(' ','')=='outputfile':
                                wb[w].cell(line_of_case,v[1]).value=thefile.replace('/','\\').replace('\\\\','\\')+'.out'
                            if v[0].lower().replace(' ','')==arg.lower():
                                if v[2]!="[str]":
                                    try:
                                        wb[w].cell(line_of_case,v[1]).value=float(all_arg2[z])
                                    except:
                                        wb[w].cell(line_of_case,v[1]).value="*******"
                                else:
                                    wb[w].cell(line_of_case,v[1]).value=all_arg2[z]
                    #print(len(alllines))
                        z+=1
                f.close()
            if os.path.exists(thefile+'.lba.csv'):
                f=open(thefile+".lba.csv")
                alllines=f.readlines()
                if len(alllines)>1:
                    all_arg=alllines[0].split()
                    all_arg2=alllines[1].split()
                    z=0
                    for arg in all_arg:
                        for v in params:
                            if v[0].lower().replace(' ','')=='outputfile':
                                #print(thefile+'\n'+thefile.replace('//','\\'))
                                wb[w].cell(line_of_case,v[1]).value=thefile.replace('//','\\').replace('/','\\')+'.out'
                            if v[0].lower().replace(' ','')==arg.lower():
                                if v[2]!="[str]":
                                    try:
                                        wb[w].cell(line_of_case,v[1]).value=float(all_arg2[z])
                                    except:
                                        wb[w].cell(line_of_case,v[1]).value="*******"
                                else:
                                    wb[w].cell(line_of_case,v[1]).value=all_arg2[z]
                    #print(len(alllines))
                        z+=1
                f.close()
            #we create the cases here!

            #close and exit!
        #break

        try:
            wb.save(file)
        except:
            print("Error while saving, please close the file and we try once more!")
            input()
            wb.save(file)

#print(generate_one.Filename)
#generate_one.Filename = 'L_01502'
#generate_one.CreateCase()


#    #input()
