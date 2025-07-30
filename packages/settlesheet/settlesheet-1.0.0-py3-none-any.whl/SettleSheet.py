#!/usr/bin/env python3
"""
SettleSheet - Shared Expenses Calculator
Designed and shared for free use by Phil Cigan under a CC-BY-SA license
A spreadsheet tool for tracking and splitting group expenses across multiple currencies

Copyright (C) 2025 Phil Cigan
This program is free software: you can redistribute it and/or modify it under the terms of the
GNU General Public License as published by the Free Software Foundation, either version 3 of the
License, or (at your option) any later version.

This program is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
See the GNU General Public License for more details.

You should have received a copy of the GNU General Public License along with this program.
If not, see <https://www.gnu.org/licenses/>.
"""
__version__ = "1.0"

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Color
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import os


def create_optimized_settlement_formulas(ws, participants, balance_row, participants_row, 
                                       optimized_settlement_header_row, section_colors, 
                                       create_font, get_column_letter):
    """
    Create optimized settlement formulas using a sequential allocation approach.
    
    This function populates the optimized settlements matrix with formulas that implement
    a greedy algorithm to minimize the total number of transactions needed to settle
    all debts among participants.
    
    Parameters:
    -----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        The worksheet object where the formulas will be added
        
    participants : list
        List of participant names (strings)
        
    balance_row : int
        Row number containing the final balance calculations for each participant
        (positive values = person is owed money, negative values = person owes money)
        
    participants_row : int
        Row number containing the participant names in the spreadsheet
        
    optimized_settlement_header_row : int
        Row number of the header row for the optimized settlements matrix
        
    section_colors : dict
        Dictionary containing color theme information for styling cells
        
    create_font : function
        Function that creates font objects with consistent styling
        
    get_column_letter : function
        Function that converts column numbers to Excel column letters (A, B, C, etc.)
        
    Returns:
    --------
    None
        The function modifies the worksheet directly by adding formulas to cells
        
    Notes:
    ------
    The optimized settlements matrix shows the minimum number of transactions needed
    to settle all debts. Each cell [i,j] shows how much participant i should pay
    to participant j. The algorithm works by:
    1. Identifying participants with negative balances (debtors) and positive balances (creditors)
    2. Creating direct transfers from largest debtors to largest creditors
    3. Making each transfer as large as possible (greedy approach)
    4. Adjusting remaining balances and repeating until all debts are settled
    
    The matrix is populated row by row, column by column, with each formula
    accounting for previous payments made by the same person to avoid over-allocation.
    """
    n = len(participants)
    
    for i in range(n):
        from_row = optimized_settlement_header_row + 1 + i
        from_participant_cell = f'{get_column_letter(i+2)}{participants_row}'
        from_balance = f'{get_column_letter(i+2)}{balance_row}'
        
        for j in range(n):
            if i != j:  # Skip diagonal (already has "-")
                to_participant_cell = f'{get_column_letter(j+2)}{participants_row}'
                to_balance = f'{get_column_letter(j+2)}{balance_row}'
                
                cell = ws.cell(row=from_row, column=j+2)
                
                # Calculate already paid (sum of previous columns in this row)
                prev_payments = []
                for k in range(j):
                    if k != i:
                        prev_cell = f'{get_column_letter(k+2)}{from_row}'
                        prev_payments.append(f'IF(ISNUMBER({prev_cell}),MAX(0,{prev_cell}),0)')
                
                if prev_payments:
                    already_paid = f'({"+".join(prev_payments)})'
                else:
                    already_paid = "0"
                
                # Calculate already received (sum of previous rows in this column)
                prev_receipts = []
                for k in range(i):
                    if k != j:
                        prev_row = optimized_settlement_header_row + 1 + k
                        prev_cell = f'{get_column_letter(j+2)}{prev_row}'
                        prev_receipts.append(f'IF(ISNUMBER({prev_cell}),MAX(0,{prev_cell}),0)')
                
                if prev_receipts:
                    already_received = f'({"+".join(prev_receipts)})'
                else:
                    already_received = "0"
                
                # Build the formula
                formula = (
                    f'=IF(OR(ISBLANK({from_participant_cell}),ISBLANK({to_participant_cell})),0,'
                    f'IF(AND({from_balance}<0,{to_balance}>0),'
                    f'MIN(MAX(0,ABS({from_balance})-{already_paid}),'
                    f'MAX(0,{to_balance}-{already_received})),0))'
                    ## Test to make symmetric: replace above line with:
                    #f'MAX(0,{to_balance}-{already_received})),'
                    #f'IF(AND({to_balance}<0,{from_balance}>0),'
                    #f'-MIN(MAX(0,ABS({to_balance})-{already_received}),'
                    #f'MAX(0,{from_balance}-{already_paid})),0)))'
                    ##--> Note this test seems to give a different potential solution in the symmetric
                    #    cells, but they don't match the original (positive) amounts. Investigate more later.
                )
                
                cell.value = formula

def create_expense_spreadsheet(filename="shared_expenses.xlsx", 
                               participants=None, 
                               expense_rows=20, 
                               exchange_rates=None, 
                               native_currency="USD",
                               color_theme="neutral", 
                               background_color="auto",
                               exchange_rate_precision=4, 
                               include_optimized_settlements=True):
    """
    Generate a shared expenses spreadsheet with automatic calculations.
    
    This function creates a comprehensive Excel spreadsheet for tracking and splitting
    group expenses. It supports both single-currency and multi-currency scenarios,
    with automatic calculations for balances, settlements, and optimized payment plans.
    
    Parameters:
    -----------
    filename : str, optional
        Name of the output Excel file. Defaults to "shared_expenses.xlsx"
        
    participants : list, optional
        List of participant names (strings). If None, defaults to ["Person 1", "Person 2", "Person 3"]
        
    expense_rows : int, optional
        Number of rows to add for expense data entry. Defaults to 20
        
    exchange_rates : dict, optional
        Dictionary mapping currency codes to their exchange rates relative to native_currency.
        Set to None for domestic trips with no currency conversion.
        Example: {'EUR': 1.1, 'JPY': 0.0067, 'GBP': 1.27}
        
        Exchange rates should be specified as what 1 unit of the local currency equals
        in native currency. For example, if 1 EUR = 1.1 USD, use {'EUR': 1.1}.
        For reverse rates (e.g., 123.45 JPY = 1 USD), use {'JPY': 1/123.45}.
        
    native_currency : str, optional
        The base currency code for all calculations. Defaults to "USD"
        
    color_theme : str, optional
        Color theme to use for the spreadsheet styling. Options:
        - "bright": Vibrant colors with light blue background
        - "neutral": Warm, muted earth tones
        - "sleek": Professional dark taupe theme
        - "bold": High contrast with bright colors
        - "dark": Dark theme similar to code editors
        - "light": Clean white background with pastel colors
        Defaults to "neutral"
        
    background_color : str, optional
        Worksheet background color. Options:
        - "auto": Automatically selects appropriate color based on theme
        - Hex color code (e.g., "#FFFFFF")
        - Named color (e.g., "white", "black")
        Defaults to "auto"
        
    exchange_rate_precision : int, optional
        Number of decimal places to display for exchange rates. Defaults to 4
        
    include_optimized_settlements : bool, optional
        Whether to include the optimized settlements section that minimizes
        the total number of transactions needed. Defaults to True
        
    Returns:
    --------
    None
        The function creates and saves an Excel file to disk
        
    Notes:
    ------
    The generated spreadsheet includes the following sections:
    
    1. **Participants Section**: Editable participant names with color coding
    
    2. **Currency Settings** (multi-currency only): Exchange rate table with
       validation and automatic inverse rate calculations
    
    3. **Summary Sections**:
       - Total Each Paid: What each person has paid
       - Total Each Owes: What each person owes to others
       - Total Each Owed: What each person is owed by others
       - Final Balance: Net balance for each person
    
    4. **Settlements Sections**:
       - Direct Settlements: Shows exact amounts owed between each pair
       - Optimized Settlements: Minimal transactions to settle all debts
    
    5. **Expense Table**: Itemized expense tracking with automatic calculations
       for "who owes whom" amounts
    
    6. **Instructions**: User guidance for using the spreadsheet
    
    The spreadsheet uses Excel formulas throughout to ensure all calculations
    update automatically when data is entered or modified.
    
    Examples:
    ---------
    # Single currency domestic trip
    create_expense_spreadsheet(
        filename="domestic_trip.xlsx",
        participants=["John", "Sarah", "Mike"],
        color_theme="neutral"
    )
    
    # Multi-currency international trip
    create_expense_spreadsheet(
        filename="europe_trip.xlsx",
        participants=["Alice", "Bob", "Charlie"],
        exchange_rates={"EUR": 1.1, "GBP": 1.27},
        native_currency="USD",
        color_theme="dark"
    )
    """
    
    if participants is None:
        participants = ["Person 1", "Person 2", "Person 3"]
    
    # Process currency parameters
    currency_list = []
    
    if exchange_rates is None:
        multi_currency = False
    else:
        multi_currency = True
        # Create a list of currency info dictionaries from the exchange_rates dictionary
        for currency_code, rate in exchange_rates.items():
            currency_list.append({
                'code': currency_code,
                'rate': rate
            })
    
    # Initialize all row and column variables used throughout the function
    # Column positions
    if not multi_currency:
        # Single currency version
        amount_col = 3
        paid_by_col = 4
        participants_col = 5
        first_owed_col = 7  # Starting column for the "Who owes whom" columns
    else:
        # Multi-currency version
        amount_col = 3  # Amount Paid
        currency_col = 4  # Paid In dropdown
        native_amount_col = 5  # Amount in native currency (calculated)
        paid_by_col = 6
        participants_col = 7
        first_owed_col = 9  # Starting column for the "Who owes whom" columns

    # Number of expense rows to add
    #expense_rows = 20  # Add formulas for first several rows
    ### --> Now setting as a parameter

    # Row positions - will be set later but initialize here
    participants_row = None
    expense_first_row = None
    expense_last_row = None
    total_row = None
    amount_paid_row = None
    share_owes_row = None
    share_owed_row = None
    settlement_header_row = None
    settlements_rows_end = None
    optimized_settlement_header_row = None
    optimized_settlements_rows_end = None

    # Define color themes
    color_themes = {
        "bright": {
            "participant_colors": [
                "CCE5FF", "FFCCCC", "D6F5D6", "FFE6CC", "E6CCFF", "FFFFCC", "FFD9EC", "CCFFFF"
            ],
            "section_colors": {
                "header": "F0F0F0", "header_dark": "D0D0D0", "paid": "E6F2FF",
                "paid_total": "CCE5FF", "owed": "FFE6E6", "owed_to": "E6E6FF", 
                "balance": "E6FFE6", "settlements": "FFF9CC", "totals": "FFFFCC",
                "instructions": "F5F5F5", "instructions_header": "E0E0E0",
                "same_person": "CCCCCC", "border": "808080", "separator": "E0E0E0",
                "separator_dark": "B0B0B0", "dropdown": "ECF4FC", "calculated": "F0F5FC",
                "native_currency": "F0F5FC", "credits": "A0A0A0"
            },
            "currency_colors": [
                "D9E9F7", "F7D9D9", "E1F5E1", "F7EBDA", "EBDAF7", "F7F7DA", "F7E6F0", "DAF7F7"
            ],
            "background": "F9FCFF"  # Very light blue
        },
        "neutral": {
            "participant_colors": [
                "D2B48C", "8E9A90", "C1A9A9", "B8C1D3", "B9B276", "A292AA", "A9C1B8", "D4B886"
            ],
            "section_colors": {
                "header": "EDEBE6", "header_dark": "D6D1C8", "paid": "E8E1D5",
                "paid_total": "DFD8CB", "owed": "D8DBD3", "owed_to": "E2DFD8",
                "balance": "E6E0D6", "settlements": "EDE8DE", "totals": "DFD9CD",
                "instructions": "F0EFEB", "instructions_header": "E5E2DC",
                "same_person": "D6D1C8", "border": "A59E94", "separator": "E5E2DC",
                "separator_dark": "C8C2B6", "dropdown": "EEE9E0", "calculated": "F5F2EE",
                "native_currency": "F5F2EE", "credits": "A59E94"
            },
            "currency_colors": [
                "E2DAD0", "DAE0DB", "E0DADA", "DCE0E4", "E0DCD2", "DAD7DE", "D9E0DD", "E2D9D0"
            ],
            "background": "FAF8F5"  # Warm off-white
        },
        "sleek": {
            "participant_colors": [
                "536878", "4B6455", "856D4D", "705A52", "6B5876", "36454F", "4F4A45", "4B6455"
            ],
            "currency_colors": [
                "8A7F72", "7D8677", "977B60", "887169", "79708A", "5E6C75", "726D65", "6E7B70"
            ],
            "section_colors": {
                "header": "B0A89E", "header_dark": "9E9689", "paid": "A8B2BD",
                "paid_total": "96A0AB", "owed": "B8A8A8", "owed_to": "A8A8B8", 
                "balance": "A8B8AB", "settlements": "B8AD98", "totals": "ADA899",
                "instructions": "B5ADA2", "instructions_header": "A3998D",
                "same_person": "A19A92", "border": "8C8680", "separator": "A3998D",
                "separator_dark": "8A8278", "dropdown": "A8B0B8", "calculated": "AAA59E",
                "native_currency": "A69D8F", "credits": "968F86"
            },
            "background": "C0B8AA"  # Dark warm taupe - woody library feel
        },
        "bold": {
            "participant_colors": [
                "4169E1", "DC143C", "2E8B57", "FF8C00", "8A2BE2", "FF4500", "1E90FF", "32CD32"
            ],
            "section_colors": {
                "header": "E0E0E0", "header_dark": "B0B0B0", "paid": "A7C7E7",
                "paid_total": "6A9ACD", "owed": "E7A7A7", "owed_to": "BCA7E7",
                "balance": "A7E7B3", "settlements": "E7DCA7", "totals": "E7E4A7",
                "instructions": "F0F0F0", "instructions_header": "D5D5D5",
                "same_person": "C0C0C0", "border": "707070", "separator": "D5D5D5",
                "separator_dark": "A0A0A0", "dropdown": "E8EEF5", "calculated": "E5EFF7",
                "native_currency": "E5EFF7", "credits": "909090"
            },
            "currency_colors": [
                "7A97E0", "E0707C", "6FAA87", "FFAC55", "AE77E0", "FF8555", "74B6FF", "76DB76"
            ],
            "background": "EFF8FF"  # Light blue-gray, more colorful but still light
        },
        "dark": {
            "participant_colors": [
                "1A3A6E", "6E1A1A", "1A5E1A", "6E4A1A", "3A1A6E", "4A5E1A", "2A2A6E", "6E1A1A"
            ],
            "section_colors": {
                "header": "3D3D3D", "header_dark": "2A2A2A", "paid": "253545",
                "paid_total": "1F2C37", "owed": "372525", "owed_to": "252537",
                "balance": "253725", "settlements": "37352A", "totals": "37332A",
                "instructions": "303030", "instructions_header": "2A2A2A",
                "same_person": "383838", "border": "505050", "separator": "333333",
                "separator_dark": "252525", "dropdown": "2D3842", "calculated": "323232",
                "native_currency": "383838", "credits": "707070"
            },
            "currency_colors": [
                "2C3845", "453535", "354535", "453C30", "363045", "3C4530", "323245", "453535"
            ],
            "background": "202020"  # Darker background similar to code editor dark themes
        },
        "light": {
            "participant_colors": [
                "E6F2FF", "FFE6E6", "F0FFF0", "FFF5E6", "F5E6FF", "FFFFD9", "FFE6F0", "E6FFFF"
            ],
            "section_colors": {
                "header": "F8F8F8", "header_dark": "F0F0F0", "paid": "F0F8FF",
                "paid_total": "E6F2FF", "owed": "FFF0F0", "owed_to": "F0F0FF",
                "balance": "F0FFF0", "settlements": "FFFFF0", "totals": "FFFAF0",
                "instructions": "FAFAFA", "instructions_header": "F5F5F5",
                "same_person": "F0F0F0", "border": "D0D0D0", "separator": "F5F5F5",
                "separator_dark": "E8E8E8", "dropdown": "F8FAFF", "calculated": "FFFFFF",
                "native_currency": "FFFFFF", "credits": "C0C0C0"
            },
            "currency_colors": [
                "CCE0F5", "F5CCCC", "DDF5DD", "F5E8CC", "E8CCF5", "F5F5C2", "F5CCDF", "CCF5F5"
            ],
            "background": "FFFFFF"  # Pure white
        }
    }
    
    # Select the appropriate theme (default to "neutral" if not found)
    selected_theme = color_themes.get(color_theme.lower(), color_themes["neutral"])
    
    # Use the selected theme's colors
    participant_colors = selected_theme["participant_colors"]
    section_colors = selected_theme["section_colors"]
    currency_colors = selected_theme["currency_colors"]
    
    # Determine base font color based on theme
    if color_theme.lower() == "dark":
        base_font_color = "E0E0E0"  # White text for dark theme
    else:
        base_font_color = "222222"  # Very dark gray for all other themes
    
    # Process background color
    if background_color.lower() == "auto":
        # Use the theme's default background
        bg_color = selected_theme.get("background", "FFFFFF")
    else:
        # User specified a custom color
        bg_color = background_color
        # Remove '#' if present in hex color
        if bg_color.startswith('#'):
            bg_color = bg_color[1:]
    
    # Add currency colors to each currency in the list
    for i, curr in enumerate(currency_list):
        curr['color'] = currency_colors[i % len(currency_colors)]
    
    # Ensure we have enough colors for all participants
    while len(participant_colors) < len(participants):
        participant_colors.extend(participant_colors)
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Shared Expenses"
    
    # Helper function to create fonts while preserving base font color
    def create_font(bold=None, size=None, italic=None, underline=None, color=None, name=None):
        return Font(
            bold=bold,
            size=size,
            italic=italic,
            underline=underline,
            color=color if color else base_font_color,
            name=name
        )
    
    # Helper function to create a safe cell reference
    def safe_cell_reference(cell_ref):
        return f'IF(ISBLANK({cell_ref}), "", {cell_ref})'
    
    # Set the default font color and style
    wb._named_styles['Normal'].font = create_font()

    # Apply background color to the entire worksheet
    ws.sheet_properties.tabColor = bg_color
    
    # Set default column width for the entire sheet
    ws.sheet_properties.defaultColWidth = 15
    
    # Set default fill for all cells (background color)
    default_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    wb._named_styles['Normal'].fill = default_fill
    
    # We'll apply this fill as needed to cells that don't have other specific fills
    for row_idx in range(1, 1000):  # Adjust range as needed
        row = ws.row_dimensions[row_idx]
        row.fill = default_fill
    
    # Set global column widths - Set this once and don't override later
    column_widths = {
        'A': 20,     # Increased width of column A
        'B': 25,     # Description
        'C': 15,     # Amount paid
        'D': 12,     # Paid in
        'E': 15,     # Native amount
        'F': 20,     # Paid by
        'G': 20,     # Participants
        'H': 15,     # Separator column before who-owes-whom
    }
    
    # Apply column widths
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Ensure the last participant column (which will contain the Grand Group Total) is wide enough
    last_col = get_column_letter(len(participants) + 1)
    ws.column_dimensions[last_col].width = 15
    
    # Row counter to keep track of current row
    current_row = 1
    
    # ===== Title and description =====
    ws.cell(row=current_row, column=1).value = "Shared Expenses Calculator"
    ws.cell(row=current_row, column=1).font = create_font(size=16, bold=True)
    ws.cell(row=current_row, column=1).fill = default_fill
    current_row += 1
    
    ws.cell(row=current_row, column=1).value = "Track and split group expenses automatically"
    ws.cell(row=current_row, column=1).font = create_font()
    ws.cell(row=current_row, column=1).fill = default_fill
    current_row += 1
    
    # ===== Participants section =====
    current_row += 1  # Add space
    
    ws.cell(row=current_row, column=1).value = "Participants"
    ws.cell(row=current_row, column=1).font = create_font(bold=True)
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color=section_colors["header_dark"], end_color=section_colors["header_dark"], fill_type="solid")
    
    # Add participants with unique colors
    for i, participant in enumerate(participants):
        col = i + 2  # 2=B, 3=C, etc.
        color = participant_colors[i]
        
        ws.cell(row=current_row, column=col).value = participant
        ws.cell(row=current_row, column=col).font = create_font(bold=True)
        ws.cell(row=current_row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    # Store the participants row for later reference
    participants_row = current_row
    current_row += 1  # Add space
    
    # ===== Currency settings (only if multi_currency is True) =====
    exchange_rate_cells = {}  # Dictionary to store cell references for exchange rates
    
    if multi_currency:
        current_row += 1  # Add space
        
        # Create a currency table
        ws.cell(row=current_row, column=1).value = "Currency Settings"
        ws.cell(row=current_row, column=1).font = create_font(bold=True, size=12)
        
        # Add Exchange Rates label split across two cells
        ws.cell(row=current_row, column=3).value = "Exchange "
        ws.cell(row=current_row, column=3).font = create_font(bold=True, size=12)
        ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='right')
        
        ws.cell(row=current_row, column=4).value = " Rates"
        ws.cell(row=current_row, column=4).font = create_font(bold=True, size=12)
        ws.cell(row=current_row, column=4).alignment = Alignment(horizontal='left')
        
        current_row += 1
        
        # Currency table headers
        ws.cell(row=current_row, column=1).value = "Currency"
        ws.cell(row=current_row, column=1).font = create_font(bold=True)
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color=section_colors["header_dark"], end_color=section_colors["header_dark"], fill_type="solid")
        
        ws.cell(row=current_row, column=2).value = "Code"
        ws.cell(row=current_row, column=2).font = create_font(bold=True)
        ws.cell(row=current_row, column=2).fill = PatternFill(start_color=section_colors["header_dark"], end_color=section_colors["header_dark"], fill_type="solid")
        
        ws.cell(row=current_row, column=3).value = f"1 unit = ? {native_currency}"
        ws.cell(row=current_row, column=3).font = create_font(bold=True)
        ws.cell(row=current_row, column=3).fill = PatternFill(start_color=section_colors["header_dark"], end_color=section_colors["header_dark"], fill_type="solid")
        
        ws.cell(row=current_row, column=4).value = f"1 {native_currency} = ? units"
        ws.cell(row=current_row, column=4).font = create_font(bold=True)
        ws.cell(row=current_row, column=4).fill = PatternFill(start_color=section_colors["header_dark"], end_color=section_colors["header_dark"], fill_type="solid")
        
        current_row += 1
        
        # Add each local currency with its rate
        currency_row_start = current_row
        for i, curr in enumerate(currency_list):
            # Currency name (e.g., "Euro")
            currency_name = {
                'AED': 'UAE Dirham',
                'AUD': 'Australian Dollar',
                'BGN': 'Bulgarian Lev',
                'BRL': 'Brazilian Real',
                'CAD': 'Canadian Dollar',
                'CHF': 'Swiss Franc',
                'CNY': 'Chinese Yuan',
                'CZK': 'Czech Koruna',
                'DKK': 'Danish Krone',
                'EUR': 'Euro',
                'GBP': 'British Pound',
                'HKD': 'Hong Kong Dollar',
                'HRK': 'Croatian Kuna',
                'HUF': 'Hungarian Forint',
                'IDR': 'Indonesian Rupiah',
                'ILS': 'Israeli Shekel',
                'INR': 'Indian Rupee',
                'ISK': 'Icelandic Króna',
                'JPY': 'Japanese Yen',
                'KRW': 'South Korean Won',
                'MXN': 'Mexican Peso',
                'MYR': 'Malaysian Ringgit',
                'NOK': 'Norwegian Krone',
                'NZD': 'New Zealand Dollar',
                'PHP': 'Philippine Peso',
                'PLN': 'Polish Złoty',
                'RON': 'Romanian Leu',
                'RUB': 'Russian Ruble',
                'SEK': 'Swedish Krona',
                'SGD': 'Singapore Dollar',
                'THB': 'Thai Baht',
                'TRY': 'Turkish Lira',
                'USD': 'US Dollar',
                'ZAR': 'South African Rand'
            }.get(curr['code'], f"Currency ({curr['code']})")  # Fallback includes code for unknown currencies
            
            ws.cell(row=current_row, column=1).value = currency_name
            ws.cell(row=current_row, column=1).font = create_font()
            ws.cell(row=current_row, column=1).fill = PatternFill(start_color=curr['color'], end_color=curr['color'], fill_type="solid")
            
            ws.cell(row=current_row, column=2).value = curr['code']
            ws.cell(row=current_row, column=2).font = create_font(bold=True)
            ws.cell(row=current_row, column=2).fill = PatternFill(start_color=curr['color'], end_color=curr['color'], fill_type="solid")
            
            ws.cell(row=current_row, column=3).value = curr['rate']
            ws.cell(row=current_row, column=3).font = create_font()
            ws.cell(row=current_row, column=3).number_format = f'#,##0.{"0" * exchange_rate_precision}'  # Dynamic precision
            ws.cell(row=current_row, column=3).fill = PatternFill(start_color=curr['color'], end_color=curr['color'], fill_type="solid")
            
            # Add inverse rate formula
            ws.cell(row=current_row, column=4).value = f'=1/C{current_row}'
            ws.cell(row=current_row, column=4).font = create_font()
            ws.cell(row=current_row, column=4).number_format = f'#,##0.{"0" * exchange_rate_precision}'  # Dynamic precision
            ws.cell(row=current_row, column=4).fill = PatternFill(start_color=curr['color'], end_color=curr['color'], fill_type="solid")
            
            # Add validation to ensure positive exchange rates
            dv = DataValidation(type="decimal", operator="greaterThan", formula1="0")
            dv.error = "Exchange rate must be greater than 0"
            dv.errorTitle = "Invalid Exchange Rate"
            ws.add_data_validation(dv)
            dv.add(ws.cell(row=current_row, column=3))
            
            # Store cell reference for this currency's exchange rate
            exchange_rate_cells[curr['code']] = f"$C${current_row}"
            current_row += 1
        
        # Add native currency row
        ws.cell(row=current_row, column=1).value = "Native Currency"
        ws.cell(row=current_row, column=1).font = create_font()
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color=section_colors["native_currency"], end_color=section_colors["native_currency"], fill_type="solid")
        
        ws.cell(row=current_row, column=2).value = native_currency
        ws.cell(row=current_row, column=2).font = create_font(bold=True)
        ws.cell(row=current_row, column=2).fill = PatternFill(start_color=section_colors["native_currency"], end_color=section_colors["native_currency"], fill_type="solid")
        
        ws.cell(row=current_row, column=3).value = 1.0
        ws.cell(row=current_row, column=3).font = create_font()
        ws.cell(row=current_row, column=3).number_format = '#,##0.0000'  # Add number formatting
        ws.cell(row=current_row, column=3).fill = PatternFill(start_color=section_colors["native_currency"], end_color=section_colors["native_currency"], fill_type="solid")
        
        ws.cell(row=current_row, column=4).value = 1.0
        ws.cell(row=current_row, column=4).font = create_font()
        ws.cell(row=current_row, column=4).number_format = '#,##0.0000'  # Add number formatting
        ws.cell(row=current_row, column=4).fill = PatternFill(start_color=section_colors["native_currency"], end_color=section_colors["native_currency"], fill_type="solid")
        
        currency_row_end = current_row
        current_row += 1
        
        # Add currency conversion link in a single cell
        # Define the hyperlink color based on theme
        if color_theme.lower() == "dark":
            # For dark theme - brighter blue with high contrast
            hyperlink_color = "6699FF"  # Bright medium blue
        else:
            # For all light themes - softer blue that's still visible on sleek theme
            hyperlink_color = "2E5CB8"  # Medium navy blue
        link_text = f"Check for current exchange rates: https://www.xe.com/currencyconverter/"
        hyperlink_cell = ws.cell(row=current_row, column=1)
        hyperlink_cell.hyperlink = "https://www.xe.com/currencyconverter/"
        #hyperlink_cell.value = link_text
        # Add the hyperlink formula instead of using the hyperlink property, to fix how it appears in some software
        hyperlink_cell.value = '=HYPERLINK("https://www.xe.com/currencyconverter/", "Check for current exchange rates: https://www.xe.com/currencyconverter/")'
        hyperlink_cell.font = create_font(color=hyperlink_color, underline="single")
        # Add this style setting which can help in some Excel versions
        ws.cell(row=current_row, column=1).style = "Hyperlink"
        ws.cell(row=current_row, column=1).font = Font(color=hyperlink_color, underline="single")
        
        current_row += 1
    
    # ===== Add dark gray separator line =====
    separator_row = current_row
    for col in range(1, 30):  # Extend across all potentially used columns
        ws.cell(row=separator_row, column=col).border = Border(
            bottom=Side(style='medium', color='808080')
        )
    current_row += 2  # Add two lines after separator 
    
    # Define a function to add borders to a section
    def add_section_border(start_row, end_row, start_col, end_col, color):
        # Create the border style once
        medium_border = Side(style='medium', color=color)
        
        # Get the range of cells
        for cell in ws[f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"]:
            for c in cell:
                current_border = c.border if c.border else Border()
                
                # Determine if cell is on an edge
                is_top = c.row == start_row
                is_bottom = c.row == end_row
                is_left = c.column == start_col
                is_right = c.column == end_col
                
                # Apply borders only where needed
                c.border = Border(
                    top=medium_border if is_top else current_border.top,
                    bottom=medium_border if is_bottom else current_border.bottom,
                    left=medium_border if is_left else current_border.left,
                    right=medium_border if is_right else current_border.right
                )
    
    # Create a thin border style for data cells - adjusted for dark mode
    if color_theme.lower() == "dark":
        # For dark mode, use darker gray borders that are visible but not distracting
        thin_border = Border(
            left=Side(style='thin', color='404040'),  # Darker gray for dark mode
            right=Side(style='thin', color='404040'),
            top=Side(style='thin', color='404040'),
            bottom=Side(style='thin', color='404040')
        )
    else:
        # For light modes, keep the standard light gray borders
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
    
    # Function to apply thin borders to a range of cells
    def apply_thin_borders(start_row, end_row, start_col, end_col):
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                # Don't override existing borders
                current_border = ws.cell(row=row, column=col).border
                # Check more safely to avoid NoneType errors
                if (not current_border or 
                    not hasattr(current_border, 'left') or not hasattr(current_border, 'right') or
                    not hasattr(current_border, 'top') or not hasattr(current_border, 'bottom') or
                    (not current_border.left or not current_border.left.style) and
                    (not current_border.right or not current_border.right.style) and
                    (not current_border.top or not current_border.top.style) and
                    (not current_border.bottom or not current_border.bottom.style)):
                    ws.cell(row=row, column=col).border = thin_border
    
    # ===== Total Paid section =====
    total_paid_start_row = current_row
    ws.cell(row=current_row, column=1).value = "Total Each Paid"
    ws.cell(row=current_row, column=1).font = create_font(bold=True, size=12)
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color=section_colors["header_dark"], end_color=section_colors["header_dark"], fill_type="solid")
    
    # Color the entire header row
    for col in range(2, len(participants) + 2):
        ws.cell(row=current_row, column=col).fill = PatternFill(start_color=section_colors["header_dark"], end_color=section_colors["header_dark"], fill_type="solid")
    
    current_row += 1
    
    # Fill the Total Paid section background
    for row in range(current_row, current_row + 3):
        for col in range(1, len(participants) + 2):
            if col == 1:  # First column
                ws.cell(row=row, column=col).fill = PatternFill(start_color=section_colors["paid"], end_color=section_colors["paid"], fill_type="solid")
            else:
                # Skip participant cells which will get their own color
                pass
    
    ws.cell(row=current_row, column=1).value = "Participant"
    ws.cell(row=current_row, column=1).font = create_font()
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color=section_colors["paid"], end_color=section_colors["paid"], fill_type="solid")
    
    # Get the participant name from participant row
    for i, participant in enumerate(participants):
        col = i + 2  # 2=B, 3=C, etc.
        participant_col = get_column_letter(col)
        color = participant_colors[i]
        
        # Use safe reference for participant name
        ws.cell(row=current_row, column=col).value = f'={safe_cell_reference(f"{participant_col}{participants_row}")}'
        ws.cell(row=current_row, column=col).font = create_font(bold=True)
        ws.cell(row=current_row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.cell(row=current_row, column=col).number_format = '#,##0.00'  # Add number formatting
    
    current_row += 1
    
    ws.cell(row=current_row, column=1).value = "Amount Paid"
    ws.cell(row=current_row, column=1).font = create_font()
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color=section_colors["paid"], end_color=section_colors["paid"], fill_type="solid")
    
    # Store for later reference
    amount_paid_row = current_row
    
    current_row += 1
    
    # Add Grand Total row with darker background
    ws.cell(row=current_row, column=1).value = "Grand group total:"
    ws.cell(row=current_row, column=1).font = create_font(bold=True)
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color=section_colors["paid_total"], end_color=section_colors["paid_total"], fill_type="solid")
    
    # Fill the entire total row with darker blue
    for col in range(2, len(participants) + 2):
        ws.cell(row=current_row, column=col).fill = PatternFill(start_color=section_colors["paid_total"], end_color=section_colors["paid_total"], fill_type="solid")
    
    # Sum all amounts paid in the last column
    last_col = len(participants) + 1
    last_participant_col = get_column_letter(last_col)
    
    # Update to use SUMIF to only sum non-blank cells
    ws.cell(row=current_row, column=last_col).value = f'=SUMIFS(B{amount_paid_row}:{last_participant_col}{amount_paid_row},B{participants_row}:{last_participant_col}{participants_row},"<>")'
    
    # Ensure two-decimal formatting for Grand Total
    ws.cell(row=current_row, column=last_col).number_format = '#,##0.00'
    if native_currency:
        ws.cell(row=current_row, column=last_col).number_format = f'#,##0.00 "{native_currency}"'
    
    ws.cell(row=current_row, column=last_col).font = create_font(bold=True)
    
    total_paid_end_row = current_row
    grand_total_cell = f'{last_participant_col}{total_paid_end_row}'
    
    # Apply thin borders first (participants row already has distinct styling)
    apply_thin_borders(total_paid_start_row+1, total_paid_end_row, 1, len(participants)+1)
    # Then add section border
    add_section_border(total_paid_start_row, total_paid_end_row, 1, len(participants) + 1, section_colors["border"])
    
    current_row += 2  # Add space between sections
    
    #===== Total Each Owes section (← this is the sum of what each person owes others) =====
    total_owes_start_row = current_row
    ws.cell(row=current_row, column=1).value = "Total Each Owes"
    ws.cell(row=current_row, column=1).font = create_font(bold=True, size=12)
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color=section_colors["header_dark"], end_color=section_colors["header_dark"], fill_type="solid")
    
    # Add clarifying note
    ws.cell(row=current_row, column=2).value = "(← this is the sum of what each specific person owes others)"
    ws.cell(row=current_row, column=2).font = create_font(italic=True)
    
    # Color the entire header row
    for col in range(2, len(participants) + 2):
        ws.cell(row=current_row, column=col).fill = PatternFill(start_color=section_colors["header_dark"], end_color=section_colors["header_dark"], fill_type="solid")
    
    current_row += 1
    
    # Fill the Share Owed section background
    for row in range(current_row, current_row + 3):  # Now 3 rows (added actual share)
        for col in range(1, len(participants) + 2):
            if col == 1:  # First column
                ws.cell(row=row, column=col).fill = PatternFill(start_color=section_colors["owed"], end_color=section_colors["owed"], fill_type="solid")
            else:
                # Skip participant cells which will get their own color
                pass
    
    ws.cell(row=current_row, column=1).value = "Participant"
    ws.cell(row=current_row, column=1).font = create_font()
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color=section_colors["owed"], end_color=section_colors["owed"], fill_type="solid")
    
    # Get the participant name from participant row
    for i, participant in enumerate(participants):
        col = i + 2  # 2=B, 3=C, etc.
        participant_col = get_column_letter(col)
        color = participant_colors[i]
        
        # Use safe reference for participant name
        ws.cell(row=current_row, column=col).value = f'={safe_cell_reference(f"{participant_col}{participants_row}")}'
        ws.cell(row=current_row, column=col).font = create_font(bold=True)
        ws.cell(row=current_row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.cell(row=current_row, column=col).number_format = '#,##0.00'  # Add number formatting
    
    current_row += 1
    
    ws.cell(row=current_row, column=1).value = "Share Owes"
    ws.cell(row=current_row, column=1).font = create_font()
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color=section_colors["owed"], end_color=section_colors["owed"], fill_type="solid")
    
    # Store for later reference
    share_owes_row = current_row
    
    current_row += 1
    
    ws.cell(row=current_row, column=1).value = "Or, equal split:"
    ws.cell(row=current_row, column=1).font = create_font()
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color=section_colors["owed"], end_color=section_colors["owed"], fill_type="solid")
    
    # Add equal split calculation
    # This is Grand Total / number of participants MINUS what this person has paid
    for i, participant in enumerate(participants):
        col = i + 2  # 2=B, 3=C, etc.
        participant_col = get_column_letter(col)
        participant_cell = f'{participant_col}{participants_row}'
        
        # Calculate the equal split, handling blank participants
        formula = f'=IF(ISBLANK({participant_cell}),"",{grand_total_cell}/' + \
                 f'COUNTIF(B{participants_row}:{get_column_letter(1+len(participants))}{participants_row},"<>")-' + \
                 f'IF(ISBLANK({participant_col}{amount_paid_row}),0,{participant_col}{amount_paid_row}))'
        
        ws.cell(row=current_row, column=col).value = formula
        ws.cell(row=current_row, column=col).fill = PatternFill(start_color=section_colors["owed"], end_color=section_colors["owed"], fill_type="solid")
        ws.cell(row=current_row, column=col).font = create_font()  # Add font formatting for equal split values
        
        # Ensure proper two-decimal formatting
        ws.cell(row=current_row, column=col).number_format = '#,##0.00'
        if native_currency:
            ws.cell(row=current_row, column=col).number_format = '#,##0.00'  # Remove currency code for equal split
    
    equal_split_row = current_row
    total_owes_end_row = current_row
    
    # Apply thin borders first
    apply_thin_borders(total_owes_start_row+1, total_owes_end_row, 1, len(participants)+1)
    # Then add section border
    add_section_border(total_owes_start_row, total_owes_end_row, 1, len(participants) + 1, section_colors["border"])
    
    current_row += 2  # Add space between sections
    
    # ===== Total Each Owed section (← this is the sum of what others owe each specific person) =====
    total_owed_start_row = current_row
    ws.cell(row=current_row, column=1).value = "Total Each Owed"
    ws.cell(row=current_row, column=1).font = create_font(bold=True, size=12)
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color=section_colors["header_dark"], end_color=section_colors["header_dark"], fill_type="solid")
    
    # Add clarifying note
    ws.cell(row=current_row, column=2).value = "(← this is the sum of what others owe each specific person)"
    ws.cell(row=current_row, column=2).font = create_font(italic=True)
    
    # Color the entire header row
    for col in range(2, len(participants) + 2):
        ws.cell(row=current_row, column=col).fill = PatternFill(start_color=section_colors["header_dark"], end_color=section_colors["header_dark"], fill_type="solid")
    
    current_row += 1
    
    # Fill the Total Owed section background
    for row in range(current_row, current_row + 2):  # Only 2 rows without equal split
        for col in range(1, len(participants) + 2):
            if col == 1:  # First column
                ws.cell(row=row, column=col).fill = PatternFill(start_color=section_colors["owed_to"], end_color=section_colors["owed_to"], fill_type="solid")
            else:
                # Skip participant cells which will get their own color
                pass
    
    ws.cell(row=current_row, column=1).value = "Participant"
    ws.cell(row=current_row, column=1).font = create_font()
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color=section_colors["owed_to"], end_color=section_colors["owed_to"], fill_type="solid")
    
    # Get the participant name from participant row
    for i, participant in enumerate(participants):
        col = i + 2  # 2=B, 3=C, etc.
        participant_col = get_column_letter(col)
        color = participant_colors[i]
        
        # Use safe reference for participant name
        ws.cell(row=current_row, column=col).value = f'={safe_cell_reference(f"{participant_col}{participants_row}")}'
        ws.cell(row=current_row, column=col).font = create_font(bold=True)
        ws.cell(row=current_row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.cell(row=current_row, column=col).number_format = '#,##0.00'  # Add number formatting
    
    current_row += 1
    
    ws.cell(row=current_row, column=1).value = "Share Owed"
    ws.cell(row=current_row, column=1).font = create_font()
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color=section_colors["owed_to"], end_color=section_colors["owed_to"], fill_type="solid")
    
    # Store for later reference
    share_owed_row = current_row
    
    current_row += 1
    
    total_owed_end_row = current_row - 1  # End row is now one less since we removed equal split row
    
    # Apply thin borders first
    apply_thin_borders(total_owed_start_row+1, total_owed_end_row, 1, len(participants)+1)
    # Then add border around Total Owed section
    add_section_border(total_owed_start_row, total_owed_end_row, 1, len(participants) + 1, section_colors["border"])
    
    current_row += 2  # Add space between sections
    
    # ===== Balance section =====
    balance_start_row = current_row
    ws.cell(row=current_row, column=1).value = "Final Balance (Total owed minus total owes)"
    ws.cell(row=current_row, column=1).font = create_font(bold=True, size=12)
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color=section_colors["header_dark"], end_color=section_colors["header_dark"], fill_type="solid")
    
    # Color the entire header row
    for col in range(2, len(participants) + 2):
        ws.cell(row=current_row, column=col).fill = PatternFill(start_color=section_colors["header_dark"], end_color=section_colors["header_dark"], fill_type="solid")
    
    current_row += 1
    
    # Fill the Balance section background
    for row in range(current_row, current_row + 2):
        for col in range(1, len(participants) + 2):
            if col == 1:  # First column
                ws.cell(row=row, column=col).fill = PatternFill(start_color=section_colors["balance"], end_color=section_colors["balance"], fill_type="solid")
            else:
                # Skip participant cells which will get their own color
                pass
    
    ws.cell(row=current_row, column=1).value = "Participant"
    ws.cell(row=current_row, column=1).font = create_font()
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color=section_colors["balance"], end_color=section_colors["balance"], fill_type="solid")
    
    # Get the participant name from participant row
    for i, participant in enumerate(participants):
        col = i + 2  # 2=B, 3=C, etc.
        participant_col = get_column_letter(col)
        color = participant_colors[i]
        
        # Use safe reference for participant name
        ws.cell(row=current_row, column=col).value = f'={safe_cell_reference(f"{participant_col}{participants_row}")}'
        ws.cell(row=current_row, column=col).font = create_font(bold=True)
        ws.cell(row=current_row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.cell(row=current_row, column=col).number_format = '#,##0.00'  # Add number formatting
    
    current_row += 1
    
    ws.cell(row=current_row, column=1).value = "Net Balance"
    ws.cell(row=current_row, column=1).font = create_font()
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color=section_colors["balance"], end_color=section_colors["balance"], fill_type="solid")
    
    for i, participant in enumerate(participants):
        col = i + 2  # 2=B, 3=C, etc.
        participant_col = get_column_letter(col)
        participant_cell = f'{participant_col}{participants_row}'
        
        # Net balance is Share Owed MINUS Share Owes with blank participant check
        ws.cell(row=current_row, column=col).value = f'=IF(ISBLANK({participant_cell}),"",{participant_col}{share_owed_row}-{participant_col}{share_owes_row})'
        ws.cell(row=current_row, column=col).fill = PatternFill(start_color=section_colors["balance"], end_color=section_colors["balance"], fill_type="solid")
        ws.cell(row=current_row, column=col).font = create_font()  # Add font formatting for balance values
        
        # Ensure proper two-decimal formatting
        ws.cell(row=current_row, column=col).number_format = '#,##0.00'
        if native_currency:
            ws.cell(row=current_row, column=col).number_format = '#,##0.00'  # Remove currency code for net balance
    
    balance_row = current_row
    balance_end_row = current_row
    
    # Apply thin borders first
    apply_thin_borders(balance_start_row+1, balance_end_row, 1, len(participants)+1)
    # Then add border around Balance section
    add_section_border(balance_start_row, balance_end_row, 1, len(participants) + 1, section_colors["border"])
    
    current_row += 2  # Add space between sections
    
    # ===== Detailed Direct Settlements section =====
    settlements_start_row = current_row
    ws.cell(row=current_row, column=1).value = "Settlements (Who Pays Whom - direct to each)"
    ws.cell(row=current_row, column=1).font = create_font(bold=True, size=12)
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color=section_colors["header_dark"], end_color=section_colors["header_dark"], fill_type="solid")
    
    # Color the entire header row
    for col in range(2, len(participants) + 2):
        ws.cell(row=current_row, column=col).fill = PatternFill(start_color=section_colors["header_dark"], end_color=section_colors["header_dark"], fill_type="solid")
    
    current_row += 1
    
    # Create a settlements table showing who owes whom
    ws.cell(row=current_row, column=1).value = "From ⇩ / To ⇒ "
    ws.cell(row=current_row, column=1).font = create_font(bold=True)
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color=section_colors["settlements"], end_color=section_colors["settlements"], fill_type="solid")
    
    # Store the settlement header row
    settlement_header_row = current_row
    
    # Add participant names as column headers
    for i, participant in enumerate(participants):
        col = i + 2  # 2=B, 3=C, etc.
        participant_col = get_column_letter(col)
        color = participant_colors[i]
        
        ws.cell(row=current_row, column=col).value = f'={participant_col}{participants_row}'
        ws.cell(row=current_row, column=col).font = create_font(bold=True)
        ws.cell(row=current_row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.cell(row=current_row, column=col).number_format = '#,##0.00'  # Add number formatting
    
    current_row += 1
    
    # Add rows for each participant
    for i, from_participant in enumerate(participants):
        row = current_row + i
        from_col = i + 2  # 2=B, 3=C, etc.
        from_cell = f'{get_column_letter(from_col)}{participants_row}'
        color = participant_colors[i]
        
        # Add participant name in first column
        ws.cell(row=row, column=1).value = f'={from_cell}'
        ws.cell(row=row, column=1).font = create_font(bold=True)
        ws.cell(row=row, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.cell(row=row, column=1).number_format = '#,##0.00'  # Add number formatting
        
        # Calculate settlement amounts for each pair
        for j, to_participant in enumerate(participants):
            to_col = j + 2  # 2=B, 3=C, etc.
            cell = ws.cell(row=row, column=j+2)
            
            if i == j:  # Same person, no settlement needed
                cell.value = "-"
                cell.alignment = Alignment(horizontal='center')  # Center-align the dash
                cell.fill = PatternFill(start_color=section_colors["same_person"], end_color=section_colors["same_person"], fill_type="solid")
                cell.font = create_font()  # Use base font color for dash
                continue
            
            # This will be set later based on the expense calculations
            cell.fill = PatternFill(start_color=section_colors["settlements"], end_color=section_colors["settlements"], fill_type="solid")
            cell.font = create_font()  # Use base font color for settlement amounts
            cell.number_format = '#,##0.00'  # Use simpler format for compatibility
    
    settlements_rows_end = current_row + len(participants) - 1
    
    # Add note explaining the settlements table 
    settlements_end_row = settlements_rows_end  # Border ends before the note 
    
    # Apply thin borders first
    apply_thin_borders(settlement_header_row, settlements_end_row, 1, len(participants)+1)
    # Then add border around Settlements section 
    add_section_border(settlements_start_row, settlements_end_row, 1, len(participants) + 1, section_colors["border"])
    
    # Add the note after adding the border (so it's outside the border)
    current_row = settlements_rows_end + 1  # Update current_row for the note
    ws.cell(row=current_row, column=1).value = "Note: The amounts here show how much each person (row) should pay to each other person (column) to resolve what they directly owe each other."
    ws.cell(row=current_row, column=1).font = create_font(italic=True)
    # Note no longer has background color
    
    current_row += 2  # Add space between sections
    
    
    
    # ===== Optimized Settlements section (using greedy method) =====
    if include_optimized_settlements:
        optimized_settlements_start_row = current_row
        ws.cell(row=current_row, column=1).value = "Optimized Settlements (minimal transfers)"
        ws.cell(row=current_row, column=1).font = create_font(bold=True, size=12)
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color=section_colors["header_dark"], end_color=section_colors["header_dark"], fill_type="solid")
        
        # Color the entire header row
        for col in range(2, len(participants) + 2):
            ws.cell(row=current_row, column=col).fill = PatternFill(start_color=section_colors["header_dark"], end_color=section_colors["header_dark"], fill_type="solid")
        
        current_row += 1
        
        # Create a settlements table showing who owes whom
        ws.cell(row=current_row, column=1).value = "From ⇩ / To ⇒ "
        ws.cell(row=current_row, column=1).font = create_font(bold=True)
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color=section_colors["settlements"], end_color=section_colors["settlements"], fill_type="solid")
        
        # Store the optimized settlement header row
        optimized_settlement_header_row = current_row
        
        # Add participant names as column headers
        for i, participant in enumerate(participants):
            col = i + 2  # 2=B, 3=C, etc.
            participant_col = get_column_letter(col)
            color = participant_colors[i]
            
            ws.cell(row=current_row, column=col).value = f'={participant_col}{participants_row}'
            ws.cell(row=current_row, column=col).font = create_font(bold=True)
            ws.cell(row=current_row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws.cell(row=current_row, column=col).number_format = '#,##0.00'  # Add number formatting
        
        current_row += 1
        
        # Add rows for each participant
        for i, from_participant in enumerate(participants):
            row = current_row + i
            from_col = i + 2  # 2=B, 3=C, etc.
            from_cell = f'{get_column_letter(from_col)}{participants_row}'
            color = participant_colors[i]
            
            # Add participant name in first column
            ws.cell(row=row, column=1).value = f'={from_cell}'
            ws.cell(row=row, column=1).font = create_font(bold=True)
            ws.cell(row=row, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws.cell(row=row, column=1).number_format = '#,##0.00'  # Add number formatting
            
            # Create the cells for this row
            for j, to_participant in enumerate(participants):
                to_col = j + 2  # 2=B, 3=C, etc.
                cell = ws.cell(row=row, column=to_col)
                
                if i == j:  # Same person, no settlement needed
                    cell.value = "-"
                    cell.alignment = Alignment(horizontal='center')  # Center-align the dash
                    cell.fill = PatternFill(start_color=section_colors["same_person"], end_color=section_colors["same_person"], fill_type="solid")
                    cell.font = create_font()  # Use base font color for dash
                else:
                    # Just set basic properties - the formula will be added by create_optimized_settlement_formulas
                    cell.fill = PatternFill(start_color=section_colors["settlements"], end_color=section_colors["settlements"], fill_type="solid")
                    cell.font = create_font()  # Use base font color for settlement amounts
                    cell.number_format = '#,##0.00'  # Use simpler format for compatibility
        
        # Call the function to populate the formulas
        create_optimized_settlement_formulas( ws, participants, balance_row, participants_row,
            optimized_settlement_header_row, section_colors, create_font, get_column_letter )
        
        optimized_settlements_rows_end = current_row + len(participants) - 1
        optimized_settlements_end_row = optimized_settlements_rows_end
        
        # Apply thin borders first
        apply_thin_borders(optimized_settlement_header_row, optimized_settlements_end_row, 1, len(participants)+1)
        # Then add border around Optimized Settlements section
        add_section_border(optimized_settlements_start_row, optimized_settlements_end_row, 1, len(participants) + 1, section_colors["border"])
        
        # Add the note after adding the border (so it's outside the border)
        current_row = optimized_settlements_rows_end + 1  # Update current_row for the note
        ws.cell(row=current_row, column=1).value = "Note: The amounts in this table show only the minimum payments needed to settle all debts, reducing the total number of transactions."
        ws.cell(row=current_row, column=1).font = create_font(italic=True)
        # Note no longer has background color
        
        current_row += 2  # Add space between sections
    
    # Add dark gray separator line before expense table
    separator_row = current_row
    for col in range(1, 30):  # Extend across all potentially used columns
        ws.cell(row=separator_row, column=col).border = Border(
            bottom=Side(style='medium', color='808080')
        )
    current_row += 2  # Add two lines after separator
    
    # ===== Itemized Expense Table =====
    # Position is now after both settlement sections
    expense_start_row = current_row
    
    # Set up headers for expense table
    if not multi_currency:
        # Single currency version
        headers = ["Date", "Description", "Amount", "Paid By", "Participants", ""]
    else:
        # Multi-currency version
        headers = ["Date", "Description", "Amount Paid", "Paid In", f"Amount ({native_currency})", 
                  "Paid By", "Participants", ""]
    
    # Set basic headers
    for i, header in enumerate(headers):
        col = i + 1  # 1-based index: 1=A, 2=B, etc.
        # Only apply header dark fill up to Participants column, not to the blank separator column
        if col <= participants_col:
            ws.cell(row=expense_start_row, column=col).value = header
            ws.cell(row=expense_start_row, column=col).font = create_font(bold=True)
            ws.cell(row=expense_start_row, column=col).fill = PatternFill(start_color=section_colors["header_dark"], end_color=section_colors["header_dark"], fill_type="solid")
        else:
            ws.cell(row=expense_start_row, column=col).value = header
            ws.cell(row=expense_start_row, column=col).font = create_font(bold=True)
    
    # Add "Who owes whom" headers - one column for each possible pair of participants
    for i, from_participant in enumerate(participants):
        for j, to_participant in enumerate(participants):
            if i != j:  # Skip same person
                col = first_owed_col + i * len(participants) + j
                ws.column_dimensions[get_column_letter(col)].width = 15
                
                header_text = f"{participants[i]} owes {participants[j]}"
                ws.cell(row=expense_start_row, column=col).value = header_text
                ws.cell(row=expense_start_row, column=col).font = create_font(bold=True)
                ws.cell(row=expense_start_row, column=col).fill = PatternFill(start_color=section_colors["header_dark"], end_color=section_colors["header_dark"], fill_type="solid")
    
    # Add comments to explain the format for various columns
    paid_by_comment = Comment("Enter names separated by commas (e.g., 'John, Sarah')", "Expense Calculator")
    ws.cell(row=expense_start_row, column=paid_by_col).comment = paid_by_comment
    
    participants_comment = Comment("Enter 'All' or names separated by commas (e.g., 'John, Sarah')", "Expense Calculator")
    ws.cell(row=expense_start_row, column=participants_col).comment = participants_comment
    
    if multi_currency:
        # Add comment for Amount Paid column
        amount_comment = Comment(f"Enter the expense amount in whatever currency it was paid in", "Expense Calculator")
        ws.cell(row=expense_start_row, column=amount_col).comment = amount_comment
        
        # Add comment for Native Amount column
        native_comment = Comment(f"This is calculated automatically - do not edit", "Expense Calculator")
        ws.cell(row=expense_start_row, column=native_amount_col).comment = native_comment
    
    # Store for later reference
    expense_first_row = expense_start_row + 1  # First data row
    expense_last_row = expense_first_row + expense_rows - 1  # Last data row
    total_row = expense_first_row + expense_rows  # Total row is after the last data row
    
    # Create currency dropdown if using multi-currency
    if multi_currency:
        # Create a list of all currency codes including native currency
        all_currencies = [curr['code'] for curr in currency_list] + [native_currency]
        currency_codes_string = ",".join(all_currencies)
        
        # Create a data validation with a dropdown menu for currency selection
        currency_validation = DataValidation(
            type="list",
            formula1=f'"{currency_codes_string}"',
            allow_blank=True
        )
        currency_validation.error = "Please select a valid currency"
        currency_validation.errorTitle = "Invalid Currency"
        currency_validation.prompt = "Select currency"
        currency_validation.promptTitle = "Currency Selection"
        ws.add_data_validation(currency_validation)
    
    # Add formulas for expense rows
    for row_idx in range(expense_rows):
        row = expense_first_row + row_idx
        
        # Set font color for all cells in the expense table
        for col in range(1, participants_col + 1):
            ws.cell(row=row, column=col).font = create_font()
        
        # Set font color and ensure background color for "who owes whom" columns
        for i, from_participant in enumerate(participants):
            for j, to_participant in enumerate(participants):
                if i != j:  # Skip same person
                    col = first_owed_col + i * len(participants) + j
                    cell = ws.cell(row=row, column=col)
                    cell.font = create_font()
                    if not cell.fill or cell.fill.fill_type == 'none':
                        cell.fill = default_fill
        
        if not multi_currency:
            # Format Amount column for single currency
            ws.cell(row=row, column=amount_col).number_format = '#,##0.00'
            
        else:
            # Format Amount Paid column
            ws.cell(row=row, column=amount_col).number_format = '#,##0.00'
            
            # Add currency dropdown
            currency_validation.add(f'{get_column_letter(currency_col)}{row}')
            
            # Set initial currency and color
            initial_currency = currency_list[0]['code'] if currency_list else native_currency
            ws.cell(row=row, column=currency_col).value = initial_currency
            
            # Create conditional formatting for currency colors
            for curr in currency_list:
                dxf = DifferentialStyle(fill=PatternFill(bgColor=curr['color']))
                rule = Rule(type="cellIs", operator="equal", formula=[f'"{curr["code"]}"'],
                          dxf=dxf)
                ws.conditional_formatting.add(f'{get_column_letter(currency_col)}{row}', rule)
            
            # Add conditional formatting for native currency
            dxf = DifferentialStyle(fill=PatternFill(bgColor=section_colors["native_currency"]))
            rule = Rule(type="cellIs", operator="equal", formula=[f'"{native_currency}"'],
                      dxf=dxf)
            ws.conditional_formatting.add(f'{get_column_letter(currency_col)}{row}', rule)
            
            # Add formula for Native Amount calculation using all possible currencies
            # This gets complex with multiple currencies, need to nest IF statements
            formula_parts = []
            
            # Add equals sign at the beginning of the formula
            formula_parts.append('=')
            
            # Check if amount is blank first
            formula_parts.append(f'IF(ISBLANK({get_column_letter(amount_col)}{row}),"",')
            
            # Add nested IF statements for each currency
            for i, curr in enumerate(currency_list):
                currency_code = curr['code']
                rate_cell = exchange_rate_cells.get(currency_code)
                if rate_cell:
                    formula_parts.append(f'IF({get_column_letter(currency_col)}{row}="{currency_code}",{get_column_letter(amount_col)}{row}*{rate_cell},')
            
            # Add native currency case
            formula_parts.append(f'IF({get_column_letter(currency_col)}{row}="{native_currency}",{get_column_letter(amount_col)}{row},"")')
            
            # Close all the nested IFs
            formula_parts.append(')'*(len(currency_list) + 1))
            
            # Combine all parts into the final formula
            native_amount_formula = ''.join(formula_parts)
            
            ws.cell(row=row, column=native_amount_col).value = native_amount_formula
            ws.cell(row=row, column=native_amount_col).number_format = '#,##0.00'  # Use simpler format
            ws.cell(row=row, column=native_amount_col).fill = PatternFill(start_color=section_colors["calculated"], end_color=section_colors["calculated"], fill_type="solid")
    
        # Add formulas for "Who owes whom" columns
        for i, from_participant in enumerate(participants):
            for j, to_participant in enumerate(participants):
                if i != j:  # Skip same person
                    col = first_owed_col + i * len(participants) + j
                    
                    # Get the participant cell references
                    from_participant_cell = f'{get_column_letter(i+2)}{participants_row}'
                    to_participant_cell = f'{get_column_letter(j+2)}{participants_row}'
                    
                    # Calculate what person[i] owes to person[j] for this expense
                    # 1. If person[j] paid and person[i] is a participant (or All), then person[i] owes their share
                    # 2. Otherwise, they owe nothing
                    # 3. If either participant is blank, then no payment is needed
                    if not multi_currency:
                        # Single currency - WITH PARTICIPANT BLANK CHECK
                        formula = (
                            f'=IF(OR(ISBLANK({from_participant_cell}),ISBLANK({to_participant_cell}),ISBLANK({get_column_letter(amount_col)}{row}),ISBLANK({get_column_letter(paid_by_col)}{row}),ISBLANK({get_column_letter(participants_col)}{row})),"",'+
                            f'IF(AND({get_column_letter(paid_by_col)}{row}={to_participant_cell},'
                            f'OR({get_column_letter(participants_col)}{row}="All",'
                            f'ISNUMBER(SEARCH("," & {from_participant_cell} & ",", "," & SUBSTITUTE({get_column_letter(participants_col)}{row}," ","") & ",")))),'
                            # Count active participants for "All" case or count specific participants
                            f'{get_column_letter(amount_col)}{row}/IF({get_column_letter(participants_col)}{row}="All",'
                            f'COUNTIF(B{participants_row}:{get_column_letter(1+len(participants))}{participants_row},"<>"),'
                            f'LEN(SUBSTITUTE({get_column_letter(participants_col)}{row}," ",""))-'
                            f'LEN(SUBSTITUTE(SUBSTITUTE({get_column_letter(participants_col)}{row}," ",""),",",""))+1),0))'
                        )
                    else:
                        # Multi-currency (using native amount) - WITH PARTICIPANT BLANK CHECK
                        formula = (
                            f'=IF(OR(ISBLANK({from_participant_cell}),ISBLANK({to_participant_cell}),ISBLANK({get_column_letter(amount_col)}{row}),ISBLANK({get_column_letter(native_amount_col)}{row}),ISBLANK({get_column_letter(paid_by_col)}{row}),ISBLANK({get_column_letter(participants_col)}{row})),"",'+
                            f'IF(AND({get_column_letter(paid_by_col)}{row}={to_participant_cell},'
                            f'OR({get_column_letter(participants_col)}{row}="All",'
                            f'ISNUMBER(SEARCH("," & {from_participant_cell} & ",", "," & SUBSTITUTE({get_column_letter(participants_col)}{row}," ","") & ",")))),'
                            # Count active participants for "All" case or count specific participants
                            f'{get_column_letter(native_amount_col)}{row}/IF({get_column_letter(participants_col)}{row}="All",'
                            f'COUNTIF(B{participants_row}:{get_column_letter(1+len(participants))}{participants_row},"<>"),'
                            f'LEN(SUBSTITUTE({get_column_letter(participants_col)}{row}," ",""))-'
                            f'LEN(SUBSTITUTE(SUBSTITUTE({get_column_letter(participants_col)}{row}," ",""),",",""))+1),0))'
                        )
                    
                    ws.cell(row=row, column=col).value = formula
                    ws.cell(row=row, column=col).font = create_font()
                    ws.cell(row=row, column=col).fill = default_fill  # Explicitly set the background color
                    ws.cell(row=row, column=col).number_format = '#,##0.00'
    
    # Update expense table references
    expense_last_row = expense_first_row + expense_rows - 1  # Last data row
    
    # Add a totals row at the bottom
    total_row = expense_first_row + expense_rows
    ws.cell(row=total_row, column=1).value = "TOTALS"
    ws.cell(row=total_row, column=1).font = create_font(bold=True)
    ws.cell(row=total_row, column=1).fill = PatternFill(start_color=section_colors["totals"], end_color=section_colors["totals"], fill_type="solid")
    
    # Add right border to Participants total cell
    ws.cell(row=total_row, column=participants_col).border = Border(
        right=Side(style='medium', color=section_colors["border"])
    )
    
    # Color the main expense columns in totals row
    for col in range(2, participants_col + 1):
        ws.cell(row=total_row, column=col).fill = PatternFill(start_color=section_colors["totals"], end_color=section_colors["totals"], fill_type="solid")
    
    # Color the who-owes-whom columns in the totals row
    for i, from_participant in enumerate(participants):
        for j, to_participant in enumerate(participants):
            if i != j:  # Skip same person
                col = first_owed_col + i * len(participants) + j
                ws.cell(row=total_row, column=col).fill = PatternFill(start_color=section_colors["totals"], end_color=section_colors["totals"], fill_type="solid")
    
    if not multi_currency:
        # Single amount column
        ws.cell(row=total_row, column=amount_col).value = f'=SUM({get_column_letter(amount_col)}{expense_first_row}:{get_column_letter(amount_col)}{expense_last_row})'
        ws.cell(row=total_row, column=amount_col).font = create_font(bold=True)
        ws.cell(row=total_row, column=amount_col).number_format = '#,##0.00'
    else:
        # For multi-currency spreadsheet, sum both Amount Paid and Native Amount
        # Skip summing Amount Paid (it's in different currencies)
        # But do sum Native Amount
        ws.cell(row=total_row, column=native_amount_col).value = f'=SUM({get_column_letter(native_amount_col)}{expense_first_row}:{get_column_letter(native_amount_col)}{expense_last_row})'
        ws.cell(row=total_row, column=native_amount_col).font = create_font(bold=True)
        ws.cell(row=total_row, column=native_amount_col).number_format = '#,##0.00'
    
    # Add totals for each "Who owes whom" column
    for i, from_participant in enumerate(participants):
        for j, to_participant in enumerate(participants):
            if i != j:  # Skip same person
                col = first_owed_col + i * len(participants) + j
                
                # Sum the column
                col_letter = get_column_letter(col)
                ws.cell(row=total_row, column=col).value = f'=SUM({col_letter}{expense_first_row}:{col_letter}{expense_last_row})'
                ws.cell(row=total_row, column=col).font = create_font(bold=True)
                ws.cell(row=total_row, column=col).fill = PatternFill(start_color=section_colors["totals"], end_color=section_colors["totals"], fill_type="solid")
                
                # Ensure proper two-decimal formatting - Use simpler format
                ws.cell(row=total_row, column=col).number_format = '#,##0.00'
    
    # Now populate the "Total Paid" formulas using direct amount sums
    for i, participant in enumerate(participants):
        col = i + 2  # 2=B, 3=C, etc.
        participant_col = get_column_letter(col)
        participant_cell = f'{participant_col}{participants_row}'
        
        # Sum of expenses where this participant is the payer
        if not multi_currency:
            # For single currency spreadsheet
            amount_col_letter = get_column_letter(amount_col)
            paid_by_col_letter = get_column_letter(paid_by_col)
            
            formula = f'=IF(ISBLANK({participant_cell}),"",' + \
                     f'SUMIFS({amount_col_letter}{expense_first_row}:{amount_col_letter}{expense_last_row},' + \
                     f'{paid_by_col_letter}{expense_first_row}:{paid_by_col_letter}{expense_last_row},{participant_cell}))'
            
            ws.cell(row=amount_paid_row, column=col).value = formula
            ws.cell(row=amount_paid_row, column=col).fill = PatternFill(start_color=section_colors["paid"], end_color=section_colors["paid"], fill_type="solid")
            ws.cell(row=amount_paid_row, column=col).font = create_font()  # Add font formatting for total paid values
            
            # Ensure proper two-decimal formatting
            ws.cell(row=amount_paid_row, column=col).number_format = '#,##0.00'
        else:
            # For multi-currency spreadsheet - use native amount
            amount_col_letter = get_column_letter(native_amount_col)
            paid_by_col_letter = get_column_letter(paid_by_col)
            
            formula = f'=IF(ISBLANK({participant_cell}),"",' + \
                     f'SUMIFS({amount_col_letter}{expense_first_row}:{amount_col_letter}{expense_last_row},' + \
                     f'{paid_by_col_letter}{expense_first_row}:{paid_by_col_letter}{expense_last_row},{participant_cell}))'
            
            ws.cell(row=amount_paid_row, column=col).value = formula
            ws.cell(row=amount_paid_row, column=col).fill = PatternFill(start_color=section_colors["paid"], end_color=section_colors["paid"], fill_type="solid")
            ws.cell(row=amount_paid_row, column=col).font = create_font()  # Add font formatting for total paid values
            
            # Ensure proper two-decimal formatting
            ws.cell(row=amount_paid_row, column=col).number_format = '#,##0.00'
    
    # Now populate the "Share Owes" formulas using the sums from the "Who owes whom" columns
    for i, participant in enumerate(participants):
        col = i + 2  # 2=B, 3=C, etc.
        participant_cell = f'{get_column_letter(col)}{participants_row}'
        
        # Skip calculations for blank participants
        formula = f'=IF(ISBLANK({participant_cell}),"",('
        
        # Calculate the total that this participant owes to all others
        # This is the sum of all "[participant] owes X" columns in the expense totals
        first = True
        for j, other_participant in enumerate(participants):
            if i != j:  # Skip self
                other_participant_cell = f'{get_column_letter(j+2)}{participants_row}'
                
                if not first:
                    formula += "+"
                first = False
                
                owes_col = first_owed_col + i * len(participants) + j
                owes_col_letter = get_column_letter(owes_col)
                
                # Only include if the other participant is not blank
                formula += f'IF(ISBLANK({other_participant_cell}),0,{owes_col_letter}{total_row})'
        
        formula += "))"
        
        if first:  # If no formula parts added (shouldn't happen), add 0
            formula = f'=IF(ISBLANK({participant_cell}),"",0)'
        
        ws.cell(row=share_owes_row, column=col).value = formula
        ws.cell(row=share_owes_row, column=col).fill = PatternFill(start_color=section_colors["owed"], end_color=section_colors["owed"], fill_type="solid")
        ws.cell(row=share_owes_row, column=col).font = create_font()  # Add font formatting for share owes values
        
        # Ensure proper two-decimal formatting
        ws.cell(row=share_owes_row, column=col).number_format = '#,##0.00'
    
    # Now populate the "Share Owed" formulas using the sums from the "Who owes whom" columns
    for i, participant in enumerate(participants):
        col = i + 2  # 2=B, 3=C, etc.
        participant_cell = f'{get_column_letter(col)}{participants_row}'
        
        # Skip calculations for blank participants
        formula = f'=IF(ISBLANK({participant_cell}),"",('
        
        # Calculate the total that this participant is owed by all others
        # This is the sum of all "X owes [participant]" columns in the expense totals
        first = True
        for j, other_participant in enumerate(participants):
            if i != j:  # Skip self
                other_participant_cell = f'{get_column_letter(j+2)}{participants_row}'
                
                if not first:
                    formula += "+"
                first = False
                
                owed_col = first_owed_col + j * len(participants) + i  # Note the j,i indexing (reverse of owes)
                owed_col_letter = get_column_letter(owed_col)
                
                # Only include if the other participant is not blank
                formula += f'IF(ISBLANK({other_participant_cell}),0,{owed_col_letter}{total_row})'
        
        formula += "))"
        
        if first:  # If no formula parts added (shouldn't happen), add 0
            formula = f'=IF(ISBLANK({participant_cell}),"",0)'
        
        ws.cell(row=share_owed_row, column=col).value = formula
        ws.cell(row=share_owed_row, column=col).fill = PatternFill(start_color=section_colors["owed_to"], end_color=section_colors["owed_to"], fill_type="solid")
        ws.cell(row=share_owed_row, column=col).font = create_font()  # Add font formatting for share owed values
        
        # Ensure proper two-decimal formatting
        ws.cell(row=share_owed_row, column=col).number_format = '#,##0.00'
    
    """
    # Now populate the Direct Settlement matrix
    # Each cell should show what one person owes to another, accounting for what the other might owe them
    for i, from_participant in enumerate(participants):
        for j, to_participant in enumerate(participants):
            if i != j:  # Skip same person
                # This is the cell in the settlements matrix
                settlement_row = settlement_header_row + 1 + i
                settlement_col = j + 2
                
                # Get the participant cell references
                from_participant_cell = f'{get_column_letter(i+2)}{participants_row}'
                to_participant_cell = f'{get_column_letter(j+2)}{participants_row}'
                
                # Find the columns in the expense totals row
                owes_col = first_owed_col + i * len(participants) + j  # What from_participant owes to_participant
                owed_col = first_owed_col + j * len(participants) + i  # What to_participant owes from_participant
                
                # NET TRANSFER CALCULATION with blank participant check:
                # If either participant is blank, show nothing
                # Positive means "row person pays column person"
                # Negative means "row person receives from column person"
                formula = f'=IF(OR(ISBLANK({from_participant_cell}),ISBLANK({to_participant_cell})),"-",' + \
                          f'{get_column_letter(owes_col)}{total_row}-{get_column_letter(owed_col)}{total_row})'
                
                ws.cell(row=settlement_row, column=settlement_col).value = formula
                ws.cell(row=settlement_row, column=settlement_col).number_format = '#,##0.00'
    """
    
    # Update current_row to point after the expense table
    current_row = total_row + 2  # Add space after expense table
    
    # ===== Instructions section =====
    instructions_start_row = current_row
    ws.cell(row=current_row, column=1).value = "INSTRUCTIONS:"
    ws.cell(row=current_row, column=1).font = create_font(bold=True, size=12)
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color=section_colors["instructions_header"], end_color=section_colors["instructions_header"], fill_type="solid")
    
    # Extend the instructions header coloring to column 7 for multi-currency sheets
    max_header_col = 7 if multi_currency else 6
    for col in range(2, max_header_col):
        ws.cell(row=current_row, column=col).fill = PatternFill(start_color=section_colors["instructions_header"], end_color=section_colors["instructions_header"], fill_type="solid")
    
    current_row += 1
    
    # Add instructions with proper formatting
    if not multi_currency:
        instructions = [
            "1. Edit the participant names in the colored cells (see Participants row)",
            "2. For each expense, enter:",
            "   - Date and description",
            "   - Amount",
            "   - Who paid for the expense (single payer per line)",# (comma-separated names if multiple people paid)",
            "   - Who participated ('All' or comma-separated names)",
            "3. The summary sections at the top will automatically update",
            "4. The 'Final Balance' section shows the net balance for each person",
            "5. The 'Settlements' table shows exactly who should pay whom and how much."
        ]
    else:
        instructions = [
            "1. Edit the participant names in the colored cells (see Participants row)",
            "2. Update the currency exchange rates to match current values (check the link for current rates)",
            "3. For each expense, enter:",
            "   - Date and description",
            "   - Amount Paid (in the original currency)",
            "   - Select the currency from the 'Paid In' dropdown",
            "   - The native amount will calculate automatically using the exchange rates",
            "   - Who paid for the expense (single payer per line)",# (comma-separated names if multiple people paid)",
            "   - Who participated ('All' or comma-separated names)",
            "4. The summary sections at the top will automatically update",
            "5. The 'Final Balance' section shows the net balance for each person",
            "6. The 'Settlements' table shows exactly who should pay whom and how much."
        ]
    
    for i, instruction in enumerate(instructions):
        row = current_row + i
        ws.cell(row=row, column=1).value = instruction
        ws.cell(row=row, column=1).font = create_font()
        
        # Apply background color to instruction rows - extend to column 7 for multi-currency
        #max_col = 7 if multi_currency else 6
        max_col = 7
        for col in range(1, max_col):
            ws.cell(row=row, column=col).fill = PatternFill(start_color=section_colors["instructions"], end_color=section_colors["instructions"], fill_type="solid")
    
    current_row += len(instructions) + 1
    
    # Add note about participant names
    ws.cell(row=current_row, column=1).value = "Note: When you rename participants, their names automatically update throughout the spreadsheet."
    ws.cell(row=current_row, column=1).font = create_font()
    
    # Extend background color to column 7 for multi-currency
    max_col = 7 if multi_currency else 6
    for col in range(1, max_col):
        ws.cell(row=current_row, column=col).fill = PatternFill(start_color=section_colors["instructions"], end_color=section_colors["instructions"], fill_type="solid")
    
    current_row += 1
    
    # Add a note about removing participants
    ws.cell(row=current_row, column=1).value = "Note: You can remove a participant by clearing their name cell. All calculations will adjust automatically."
    ws.cell(row=current_row, column=1).font = create_font()
    
    # Extend background color to column 7 for multi-currency
    max_col = 7 #if multi_currency else 6
    for col in range(1, max_col):
        ws.cell(row=current_row, column=col).fill = PatternFill(start_color=section_colors["instructions"], end_color=section_colors["instructions"], fill_type="solid")
    
    instructions_end_row = current_row
    
    # Add border around Instructions section - extend border to column 6 (or 7 for multi-currency)
    max_border_col = 6 #if multi_currency else 5
    add_section_border(instructions_start_row, instructions_end_row, 1, max_border_col, section_colors["border"])
    
    # Add credits line two lines below the instructions
    credits_row = instructions_end_row + 2
    ws.cell(row=credits_row, column=1).value = "Designed and shared for free use by Phil Cigan under a CC-BY-SA license"
    ws.cell(row=credits_row, column=1).font = create_font(color=section_colors["credits"])
    
    # Apply default font color to all headers and labels
    # Update more section headers with base font color
    section_headers = [
        (total_paid_start_row, 1),       # Total Paid section
        (total_owes_start_row, 1),       # Total Each Owes section
        (total_owes_start_row, 2),       # Total Each Owes note
        (total_owed_start_row, 1),       # Total Each Owed section
        (total_owed_start_row, 2),       # Total Each Owed note
        (balance_start_row, 1),          # Balance section
        (settlements_start_row, 1),      # Settlements section
        (instructions_start_row, 1),     # Instructions section
        (total_row, 1)                   # TOTALS row
    ]
    
    for row, col in section_headers:
        cell = ws.cell(row=row, column=col)
        existing_font = cell.font if cell.font else Font()
        cell.font = Font(
            name=existing_font.name,
            size=existing_font.size,
            bold=existing_font.bold,
            italic=existing_font.italic,
            color=base_font_color
        )
    
    # Apply label colors for consistent row headers
    label_cells = [
        (current_row-len(instructions)-3, 1, "Participant"),  # Total Paid participant row
        (amount_paid_row, 1, "Amount Paid"),
        (total_paid_end_row, 1, "Grand group total:"),
        (total_owes_start_row+1, 1, "Participant"),  # Total Each Owes participant row
        (share_owes_row, 1, "Share Owes"),
        (equal_split_row, 1, "Or, equal split:"),
        (total_owed_start_row+1, 1, "Participant"),  # Total Each Owed participant row
        (share_owed_row, 1, "Share Owed"),
        (balance_start_row+1, 1, "Participant"),  # Balance participant row
        (balance_row, 1, "Net Balance"),
        (settlement_header_row, 1, "From ⇩ / To ⇒ ")
    ]
    
    for row, col, label_text in label_cells:
        cell = ws.cell(row=row, column=col)
        if cell.value == label_text:  # Just to be sure we're changing the right cell
            existing_font = cell.font if cell.font else Font()
            cell.font = create_font(
                size=existing_font.size,
                bold=existing_font.bold,
                italic=existing_font.italic
            )
    
    # Apply default font color to all headers
    for row in range(expense_start_row, expense_start_row + 1):
        for col in range(1, participants_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.font.bold:  # It's a header
                existing_font = cell.font
                cell.font = create_font(
                    name=existing_font.name,
                    size=existing_font.size,
                    bold=existing_font.bold,
                    italic=existing_font.italic
                )
    
    # For "Who owes whom" column headers
    for i, from_participant in enumerate(participants):
        for j, to_participant in enumerate(participants):
            if i != j:  # Skip same person
                col = first_owed_col + i * len(participants) + j
                cell = ws.cell(row=expense_start_row, column=col)
                existing_font = cell.font
                cell.font = create_font(
                    name=existing_font.name,
                    size=existing_font.size,
                    bold=existing_font.bold,
                    italic=existing_font.italic
                )
    
    # Apply background to all cells in the worksheet
    # First apply to all cells in a large region that covers everything
    max_row = total_row + 50  # Go well beyond the end of our defined content
    max_col = 30  # Go well beyond our widest columns
    
    for row in range(1, max_row):
        for col in range(1, max_col):
            cell = ws.cell(row=row, column=col)
            if not cell.fill.fill_type or cell.fill.fill_type == 'none':
                cell.fill = default_fill
                # Also set the default font color for any cell that has text
                if cell.value and isinstance(cell.value, str) and not cell.font.color:
                    cell.font = create_font()
    
    # Apply thin borders to all data sections
    if multi_currency:
        # Currency Settings section
        apply_thin_borders(currency_row_start, currency_row_end, 1, 4)  # Changed from 3 to 4 to include inverse rate column
    
    # Total Paid section (participants row already has distinct styling)
    apply_thin_borders(total_paid_start_row+1, total_paid_end_row, 1, len(participants)+1)
    
    # Total Each Owes section
    apply_thin_borders(total_owes_start_row+1, total_owes_end_row, 1, len(participants)+1)
    
    # Total Each Owed section
    apply_thin_borders(total_owed_start_row+1, total_owed_end_row, 1, len(participants)+1)
    
    # Balance section
    apply_thin_borders(balance_start_row+1, balance_end_row, 1, len(participants)+1)
    
    # Settlements section
    apply_thin_borders(settlement_header_row, settlements_end_row, 1, len(participants)+1)
    
    # Expense table
    apply_thin_borders(expense_start_row, total_row, 1, participants_col)
    
    # Who owes whom columns
    for i, from_participant in enumerate(participants):
        for j, to_participant in enumerate(participants):
            if i != j:  # Skip same person
                col = first_owed_col + i * len(participants) + j
                apply_thin_borders(expense_start_row, total_row, col, col)
    
    # Now add section borders to the expense table after thin borders are applied:
    # 1. Top border around header row
    for col in range(1, participants_col + 1):
        cell = ws.cell(row=expense_start_row, column=col)
        current_border = cell.border if cell.border else Border()
        cell.border = Border(
            top=Side(style='medium', color=section_colors["border"]),
            bottom=current_border.bottom,
            left=Side(style='medium', color=section_colors["border"]) if col == 1 else current_border.left,
            right=Side(style='medium', color=section_colors["border"]) if col == participants_col else current_border.right
        )
    
    # 2. Bottom border around totals row
    for col in range(1, participants_col + 1):
        cell = ws.cell(row=total_row, column=col)
        current_border = cell.border if cell.border else Border()
        cell.border = Border(
            top=current_border.top,
            bottom=Side(style='medium', color=section_colors["border"]),
            left=Side(style='medium', color=section_colors["border"]) if col == 1 else current_border.left,
            right=Side(style='medium', color=section_colors["border"]) if col == participants_col else current_border.right
        )
    
    # 3. Left and right borders from header to totals
    for row in range(expense_start_row + 1, total_row):
        # Left border
        cell = ws.cell(row=row, column=1)
        current_border = cell.border if cell.border else Border()
        cell.border = Border(
            top=current_border.top,
            bottom=current_border.bottom,
            left=Side(style='medium', color=section_colors["border"]),
            right=current_border.right
        )
        
        # Right border (Participants column)
        cell = ws.cell(row=row, column=participants_col)
        current_border = cell.border if cell.border else Border()
        cell.border = Border(
            top=current_border.top,
            bottom=current_border.bottom,
            left=current_border.left,
            right=Side(style='medium', color=section_colors["border"])
        )
    
    # Add formulas for direct settlements
    for i, from_participant in enumerate(participants):
        for j, to_participant in enumerate(participants):
            if i != j:  # Skip same person
                # Get the participant cell references
                from_participant_cell = f'{get_column_letter(i+2)}{participants_row}'
                to_participant_cell = f'{get_column_letter(j+2)}{participants_row}'
                
                # Find the columns in the expense totals row
                owes_col = first_owed_col + i * len(participants) + j  # What from_participant owes to_participant
                owed_col = first_owed_col + j * len(participants) + i  # What to_participant owes from_participant
                
                # Calculate settlement amount (what from_participant owes to_participant)
                formula = f'=IF(OR(ISBLANK({from_participant_cell}),ISBLANK({to_participant_cell})),"-",' + \
                         f'{get_column_letter(owes_col)}{total_row}-{get_column_letter(owed_col)}{total_row})'
                
                # Set the formula in the settlements table
                row = settlement_header_row + 1 + i  # +1 to skip header row
                ws.cell(row=row, column=j+2).value = formula

    # ===== Add formulas for optimized settlements using the greedy method =====
    # The optimized settlements matrix uses a greedy approach to minimize the number of transactions
    # needed to settle all debts. It works by:
    # 1. Using the final balances, where:
    #    - Negative balance means that person owes money overall
    #    - Positive balance means that person is owed money overall
    # 2. Creating direct transfers from people who owe money to people who are owed money
    # 3. Making each transfer as large as possible (greedy approach)
    #
    # Example scenario with 4 people (Alice, Bob, Charlie, David):
    # Initial expenses:
    #   - Alice pays $60 for dinner, split among Alice, Bob, Charlie ($20 each)
    #   - Bob pays $30 for drinks, split among Bob, Charlie, David ($10 each)
    #   - Charlie pays $20 for dessert, split among all four ($5 each)
    #
    # Without any optimization (raw expenses), there would be 7 separate transfers:
    #   1. Bob pays $20 to Alice (dinner share)
    #   2. Charlie pays $20 to Alice (dinner share)
    #   3. Charlie pays $10 to Bob (drinks share)
    #   4. David pays $10 to Bob (drinks share)
    #   5. Alice pays $5 to Charlie (dessert share)
    #   6. Bob pays $5 to Charlie (dessert share)
    #   7. David pays $5 to Charlie (dessert share)
    #
    # Our standard direct settlements matrix (implemented in this spreadsheet) performs pairwise netting:
    # For each pair of people (A and B), it calculates what A owes B minus what B owes A.
    # This results in the following settlements matrix:
    #   Alice | Bob | Charlie | David
    # ------------------------------- 
    # Alice   |  -  | -$20  |  -$15   |   $0
    # Bob     | $20 |   -   |    $5   | -$10
    # Charlie | $15 |  -$5  |    -    |  -$5
    # David   |  $0 |  $10  |    $5   |   -
    #
    # (Positive means row person receives from column person; negative means row person pays column person)
    # 
    # This standard pairwise netting reduces the number to 5 transactions:
    #   1. Bob pays $20 to Alice (dinner share)
    #   2. Charlie pays $15 to Alice ($20 dinner - $5 dessert)
    #   3. Charlie pays $5 to Bob ($10 drinks - $5 dessert)
    #   4. David pays $10 to Bob (drinks share)
    #   5. David pays $5 to Charlie (dessert share)
    #
    # Final balance calculation (total paid minus fair share, or summing what each person owes vs is owed):
    #   Alice:   Paid $60 - Share  (20+0+5=)$25 = +$35 (is owed money)
    #   Bob:     Paid $30 - Share (20+10+5=)$35 =  -$5 (owes money)
    #   Charlie: Paid $20 - Share (20+10+5=)$35 = -$15 (owes money)
    #   David:   Paid  $0 - Share  (0+10+5=)$15 = -$15 (owes money)
    #
    # Optimized settlement using the greedy method requires only 3 transactions:
    #   1. David pays $15 to Alice
    #   2. Charlie pays $15 to Alice
    #   3. Bob pays $5 to Alice
    # This results in the same net flow (owes vs is owed) of money for each individual person, 
    #   but with fewer transactions. Note that this could result in some people paying other
    #   people that they do not directly owe money to, but everything balances out in the end 
    #   (i.e. Alice pays David $15, but David owes Charlie $15, so Charlie pays Alice $15 directly
    #   to resolve the overall group debt).
    #
    # The greedy algorithm works by:
    #   1. Identifying who has a positive balance (owed money) and negative balance (owes money)
    #   2. Sorting people by the absolute value of their balance
    #   3. Creating direct transfers from largest debtors to largest creditors
    #   4. Making each transfer as large as possible (the minimum of what one owes and what one is owed)
    #   5. Adjusting balances and repeating until all debts are settled
    #
    # This minimizes the total number of transactions needed to settle all debts.
    #
    # Note that the optimized settlement matrix is symmetric:
    #   - Positive numbers mean "row person pays column person"
    #   - Negative numbers mean "column person pays row person"
    #   - Each cell[A,B] = -cell[B,A]
    
    # Save the workbook
    wb.save(filename)
    print(f"Spreadsheet created and saved as {filename}")
    print(f"Full path: {os.path.abspath(filename)}")

if __name__ == "__main__":
    # Example: Domestic trip (no currency conversion)
    create_expense_spreadsheet(
        filename="domestic_trip.xlsx", 
        participants=["John", "Sarah", "Mike"],
        expense_rows=20,
        exchange_rates=None,
        color_theme="neutral",
        background_color="auto"
    )
    
    # Example: International trip with ONE foreign currency
    create_expense_spreadsheet(
        filename="international_trip.xlsx",
        participants=["John", "Sarah", "Mike", "Laura"],
        expense_rows=40,
        exchange_rates={"EUR": 1.10},
        native_currency="USD",
        color_theme="sleek",
        background_color="auto"
    )
    
    # Example: Multi-national trip with MULTIPLE foreign currencies
    create_expense_spreadsheet(
        filename="multi_country_trip.xlsx",
        participants=["Alex", "Beth", "Carlos", "Dana"],
        expense_rows=60,
        exchange_rates={"EUR": 1.10, "JPY": 1/147.23, "GBP": 1.27},  # Multiple currencies
        native_currency="USD",
        color_theme="dark",
        #background_color="#FFFFFF"  # Custom background color
    )


    # # Example: Multi-national trip with MULTIPLE foreign currencies -- showcasing all color schemes
    # for color_theme in ["bright", "neutral", "sleek", "bold", "dark", "light"]:
    #     create_expense_spreadsheet(
    #         filename=f"multi_country_trip_optimizedflow_{color_theme}.xlsx",
    #         participants=["Alex", "Beth", "Carlos", "Dana", "Edward", "Fiona", "George", "Hannah"],
    #         exchange_rates={"EUR": 0.92, "JPY": 0.0068, "GBP": 1.26, "CHF": 1.13, "KRW": 0.00076, "THB": 0.028, "AUD": 0.65, "CAD": 0.74},     # Exchange rates for each currency
    #         native_currency="USD",
    #         color_theme=color_theme,
    #     )

"""
Example usage for SettleSheet.py

Basic usage (all defaults):
---------------------------
create_expense_spreadsheet()
# - Creates "shared_expenses.xlsx" for 3 participants ("Person 1", "Person 2", "Person 3")
# - 20 expense rows, no currency conversion, neutral color theme

Single-currency, custom participants and filename:
--------------------------------------------------
create_expense_spreadsheet(
    filename="domestic_trip.xlsx",
    participants=["John", "Sarah", "Mike"],
    expense_rows=25,  # Optional: set number of expense rows
    exchange_rates=None,  # No currency conversion
    color_theme="bright"  # Optional: choose a color theme
)

International trip with one foreign currency:
---------------------------------------------
create_expense_spreadsheet(
    filename="international_trip.xlsx",
    participants=["John", "Sarah", "Mike", "Laura"],
    expense_rows=40,
    exchange_rates={"EUR": 1.10},  # 1 EUR = 1.10 USD
    native_currency="USD",         # All calculations in USD
    color_theme="sleek"
)

Multi-country trip with several currencies:
-------------------------------------------
create_expense_spreadsheet(
    filename="multi_country_trip.xlsx",
    participants=["Alex", "Beth", "Carlos", "Dana"],
    expense_rows=60,
    exchange_rates={
        "EUR": 1.10,           # 1 EUR = 1.10 USD
        "JPY": 1/147.23,       # 1 JPY = 0.0068 USD
        "GBP": 1.27            # 1 GBP = 1.27 USD
    },
    native_currency="USD",
    color_theme="dark"
)

# For more customization, you can also specify:
# - background_color (hex or "auto")
# - exchange_rate_precision (number of decimal places for rates)
# - include_optimized_settlements (True/False to include the optimized payment matrix)

# See the README for more details and advanced options.

"""

