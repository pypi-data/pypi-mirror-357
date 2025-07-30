from .base import TableHandlerBase
from openpyxl import Workbook, load_workbook


class ExcelHandler(TableHandlerBase):
    """Excel适配器（使用openpyxl）"""

    def write(self, data: List[Dict], file_path: str):
        wb = Workbook()
        ws = wb.active
        if data:
            ws.append(list(data[0].keys()))
            for row in data:
                ws.append(list(row.values()))
        wb.save(file_path)

    def read(self, file_path: str) -> List[Dict]:
        wb = load_workbook(file_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0]
        return [dict(zip(headers, row)) for row in rows[1:]]