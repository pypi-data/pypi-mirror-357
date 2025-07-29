import os
import datetime
from io import BytesIO
from typing import Any, Dict, Union, Tuple, List, Optional

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font, PatternFill, Border, Alignment, Fill
from openpyxl.worksheet.worksheet import Worksheet

from rqpattr.report import excel_templates


def blank_when_none(value) -> Optional[str]:
    return value if value is not None else ""


_highlight_style = [
    Font(color="FFFFFFFF", name="微软雅黑", size=11),
    PatternFill(patternType="solid", fgColor="409cd2"),
]

_default_style = [
    Font(color="00000000", name="微软雅黑", size=11),
    PatternFill(patternType="solid", fgColor="ffffff"),
]


class SheetSchema:
    def __init__(
            self,
            sheet_id: str = "",
            title: str = None,
            row_style: Optional[
                Dict[int, List[Union[Font, Fill, Border, Alignment]]]
            ] = None,
            column_style: Optional[
                Dict[int, List[Union[Font, Fill, Border, Alignment]]]
            ] = None,
            cell_style: Optional[
                Dict[Tuple[int, int], List[Union[Font, Fill, Border, Alignment]]]
            ] = None,
            force_date: bool = False,
    ):
        self._id = "Sheet" + str(datetime.datetime.now()) if not sheet_id else sheet_id
        self._title = title
        self._row_style = row_style
        self._column_style = column_style
        self._cell_style = cell_style
        self._bond_sheet = None
        self._force_date = force_date

    @property
    def sheet(self):
        return self._bond_sheet

    @sheet.setter
    def sheet(self, sheet: Worksheet):
        self._bond_sheet = sheet

    @property
    def id(self):
        return self._id

    @property
    def title(self):
        return self._title

    def load_data(self, **kwargs):
        raise NotImplementedError

    def apply_style(self, sheet: Worksheet):
        if self._row_style is not None:
            for col in range(sheet.max_column):
                for row, style in self._row_style.items():
                    self._assign_cell_style(sheet.cell(row + 1, col + 1), style)

        if self._column_style is not None:
            for row in range(sheet.max_row):
                for col, style in self._column_style.items():
                    self._assign_cell_style(sheet.cell(row + 1, col + 1), style)

        if self._cell_style is not None:
            for (row, col), style in self._cell_style.items():
                self._assign_cell_style(sheet.cell(row + 1, col + 1), style)

    def _write_cell(
            self, sheet: Worksheet, row: int, col: int, data: object, style: Any = None
    ):
        if pd.isna(data):
            data = ""
        cell = sheet.cell(row + 1, col + 1, data)
        if style is not None:
            cell._style = style

    @staticmethod
    def _assign_cell_style(
            cell: Cell, styles: List[Union[Font, PatternFill, Border, Alignment]]
    ):
        for style in styles:
            if isinstance(style, Font):
                cell.font = style
            elif isinstance(style, Fill):
                cell.fill = style
            elif isinstance(style, Border):
                cell.border = style
            elif isinstance(style, Alignment):
                cell.alignment = style
            else:
                raise ValueError(f"Unknown style option {style}")


class ExcelTemplate:
    def __init__(self, schemas: List[SheetSchema], template_name: str = None):
        self.out = BytesIO()
        if template_name is not None:
            template_path = os.path.abspath(excel_templates.__file__)
            template_path = template_path.replace(
                "__init__.py", template_name + ".xlsx"
            )
            self._wb = openpyxl.load_workbook(template_path)
        else:
            self._wb = Workbook()

        if not schemas:
            raise ValueError("No schemas are defined.")

        self._schemas_by_order = schemas
        self._schemas_by_id: Dict[str, SheetSchema] = {s.id: s for s in schemas}

        self._sheets: List[Worksheet] = []
        for seq, schema in enumerate(schemas):
            if seq == 0:
                self._sheets.append(self._wb.worksheets[0])
                schema.sheet = self._wb.worksheets[0]
            else:
                sheet = (
                    self._wb.create_sheet(title=schema.title)
                    if len(self._wb.worksheets) <= seq
                    else self._wb.worksheets[seq]
                )
                self._sheets.append(sheet)
                schema.sheet = sheet

    def _resolve_sheet(self, sheet_id_or_seq: Union[str, int]) -> SheetSchema:
        return (
            self._schemas_by_id[sheet_id_or_seq]
            if sheet_id_or_seq in self._schemas_by_id
            else self._schemas_by_order[sheet_id_or_seq]
        )

    def load_data(self, sheet_id_or_seq: Union[str, int], **kwargs):
        schema = self._resolve_sheet(sheet_id_or_seq)
        schema.load_data(**kwargs)

    def preload_save(self, close=True):
        self._wb.save(self.out)
        self.out.seek(0)
        if close:
            self._wb.close()

    def save(self, name: str):
        for seq, schema in enumerate(self._schemas_by_order):
            schema.apply_style(self._sheets[seq])
        self._wb.save(filename=name)


class FillByLabelSchema(SheetSchema):
    def __init__(self, label_cell_map: Dict[str, Tuple[int, int]], **kwargs):
        super().__init__(**kwargs)
        self._label_cell_map = label_cell_map

    def load_data(self, data: Dict[str, Union[str, int, float, datetime.datetime]]):
        cell_data_map = {
            self._label_cell_map[label]: value for label, value in data.items()
        }

        for (row, col), data in cell_data_map.items():
            self._write_cell(self.sheet, row, col, data)


class MultiSeriesSchema(SheetSchema):
    def __init__(self, areas: Dict[str, Tuple[int, int]], **kwargs):
        super(MultiSeriesSchema, self).__init__(**kwargs)
        self._area: Dict[str, Tuple[int, int]] = areas

    def load_data(self, area: str, data: List[List]):
        if not data or not data[0]:
            return
        (row_offset, col_offset) = self._area[area]
        first_row_styles = self._get_styles(row_offset, col_offset, len(data[0]))

        # when get first row style, we can just apply these styles to remaining rows.
        self._load_data(
            row_offset=row_offset,
            col_offset=col_offset,
            data=data,
            styles=first_row_styles,
        )

    def _load_data(
            self, row_offset: int, col_offset: int, data: List[List], styles: List
    ):
        for row_num, row_data in enumerate(data):
            for col_num, cell_data in enumerate(row_data):
                self._write_cell(
                    self.sheet,
                    row_offset + row_num,
                    col_offset + col_num,
                    cell_data,
                    styles[col_num],
                )

    def _get_styles(self, row_offset: int, col_offset: int, column_length: int) -> List:
        styles = []
        for col_num in range(column_length):
            # +1 is required because excel is starts with 1.
            cell = self.sheet.cell(row_offset + 1, col_offset + col_num + 1)
            styles.append(cell._style)
        return styles


class PABrinsonTemplate(ExcelTemplate):
    def __init__(self):
        super(PABrinsonTemplate, self).__init__(
            template_name="brinson",
            schemas=[
                FillByLabelSchema(
                    sheet_id="analysis_arguments",
                    label_cell_map={
                        "benchmark": (2, 2),
                        "start_date": (3, 2),
                        "end_date": (4, 2),
                    },
                ),
                MultiSeriesSchema(
                    sheet_id="returns_decomposition",
                    areas={"returns_decomposition": (3, 0)},
                ),
                MultiSeriesSchema(
                    sheet_id="industry_attribution",
                    areas={"industry_attribution": (3, 1)},
                ),
            ],
        )


class PAFactorTemplate(ExcelTemplate):
    def __init__(self):
        super(PAFactorTemplate, self).__init__(
            template_name="factor",
            schemas=[
                FillByLabelSchema(
                    sheet_id="analysis_arguments",
                    label_cell_map={
                        "benchmark": (2, 2),
                        "start_date": (3, 2),
                        "end_date": (4, 2),
                    },
                ),
                MultiSeriesSchema(
                    sheet_id="returns_decomposition",
                    areas={"returns_decomposition": (3, 0)},
                ),
                MultiSeriesSchema(
                    sheet_id="factor_attribution",
                    areas={"factor_attribution": (3, 1)},
                ),
                MultiSeriesSchema(
                    sheet_id="factor_exposure",
                    areas={"factor_exposure": (3, 1)},
                ),
            ],
        )


class PAFactorV2Template(ExcelTemplate):
    def __init__(self):
        super(PAFactorV2Template, self).__init__(
            template_name="factor_v2",
            schemas=[
                FillByLabelSchema(
                    sheet_id="analysis_arguments",
                    label_cell_map={
                        "benchmark": (2, 2),
                        "start_date": (3, 2),
                        "end_date": (4, 2),
                    },
                ),
                MultiSeriesSchema(
                    sheet_id="returns_decomposition",
                    areas={"returns_decomposition": (3, 0)},
                ),
                MultiSeriesSchema(
                    sheet_id="factor_attribution",
                    areas={"factor_attribution": (3, 1)},
                ),
                MultiSeriesSchema(
                    sheet_id="factor_exposure",
                    areas={"factor_exposure": (3, 1)},
                ),
            ],
        )
